import pandas as pd
import os
manpower_path = r"C:\Users\ET0001301\Pictures\Data\Manpower\ManpowerME & LAG 9th March'26.xlsm"
df_mnp = pd.read_excel(manpower_path,sheet_name='Data')
branch_path = r"C:\Users\ET0001301\Pictures\Data\Manpower\Branch wise Region.xlsx"
df_branch = pd.read_excel(branch_path,sheet_name='Sheet1')

df = df_mnp[df_mnp['Department']=='ME']
df = df.drop(columns= ['Entity','Final Location','LAG Region','ME Region','Date_Of_Joining','Department','Official_Email','ReportingManager_Name','ReportingManager_Code',	'ReportingManagersManagerName','ReportingManagersManagerCode',	'SubDepartment_Code','Role1'
])
df.columns
df.insert(df.columns.get_loc('Final Region')+1,'ME Region 2',pd.NA)
df.insert(df.columns.get_loc('Role')+1,'Role 1',pd.NA)
df.columns
# mapping = df_branch.set_index('Final Branch')['ME Region 2']

# df['ME Region 2'] = df['Final Branch'].map(mapping)

# # Check if any branch was not mapped
# if df['ME Region 2'].isna().any():
#     raise ValueError('some region are missiion')
mapping = df_branch.set_index('Final Branch')['ME Region 2']

df['ME Region 2'] = df['Final Branch'].map(mapping)

# Find branches where region is not mapped
missing_branches = df.loc[df['ME Region 2'].isna(), 'Final Branch'].unique()

if len(missing_branches) > 0:
    print("Branches with missing region mapping:")
    print(missing_branches)
    raise ValueError('Some regions are missing')

# df['ME Region 2'] = df['Final Branch'].map(df_branch.set_index('Final Branch')['ME Region 2'])
df['Role 1'] = df.apply(lambda x:x['Role'] if x['Role'] in ['BSM ME','SO ME'] else "",axis = 1)
for i in range(1,4):
    df.insert(len(df.columns),f'RM CODE {i}',pd.NA)
    df.insert(len(df.columns),f'RM NAME {i}',pd.NA)
    df.insert(len(df.columns),f'RM DESIG {i}',pd.NA)
df.columns
df['RM CODE 1'] = df['Employee_Code'].map(df_mnp.set_index('Employee_Code')['ReportingManager_Code']).fillna('#N/A')
df['RM NAME 1'] = df['RM CODE 1'].map(df_mnp.set_index('Employee_Code')['Employee_Name']).fillna('#N/A')
df['RM DESIG 1'] = df['RM CODE 1'].map(df_mnp.set_index('Employee_Code')['Role']).fillna('#N/A')


df['RM CODE 2'] = df['RM CODE 1'].map(df_mnp.set_index('Employee_Code')['ReportingManager_Code']).fillna('#N/A')
df['RM NAME 2'] = df['RM CODE 2'].map(df_mnp.set_index('Employee_Code')['Employee_Name']).fillna('#N/A')
df['RM DESIG 2'] = df['RM CODE 2'].map(df_mnp.set_index('Employee_Code')['Role']).fillna('#N/A')


df['RM CODE 3']= df['RM CODE 2'].map(df_mnp.set_index('Employee_Code')['ReportingManager_Code']).fillna('#N/A')
df['RM NAME 3']= df['RM CODE 3'].map(df_mnp.set_index('Employee_Code')['Employee_Name']).fillna('#N/A')
df['RM DESIG 3']= df['RM CODE 3'].map(df_mnp.set_index('Employee_Code')['Role']).fillna('#N/A')

wrong_emplyee = df[df['RM DESIG 1'].isin(['ABM LAG','#N/A','BM','BSM LAG','SO LAG',''])]
df = df[df['Role'].isin(['ABM ME','CBM ME','TL','BSM ME','SO ME'])]
for i in range(1,4):
    df[f'RM CODE {i}'] = df.apply(lambda x:'-' if x[f'RM DESIG {i}'] not in {'ABM ME','CBM ME','TL'} else x[f'RM CODE {i}'],axis = 1)
    df[f'RM NAME {i}'] = df.apply(lambda x:'-' if x[f'RM DESIG {i}'] not in {'ABM ME','CBM ME','TL'} else x[f'RM NAME {i}'],axis = 1)
    df[f'RM DESIG {i}'] = df.apply(lambda x:'-' if x[f'RM DESIG {i}'] not in {'ABM ME','CBM ME','TL'} else x[f'RM DESIG {i}'],axis = 1)

df.insert(len(df.columns),'TL CODE',pd.NA)
df.insert(len(df.columns),'TL NAME',pd.NA)
df.insert(len(df.columns),'ABM CODE',pd.NA)
df.insert(len(df.columns),'ABM NAME',pd.NA)
df.insert(len(df.columns),'CBM CODE',pd.NA)
df.insert(len(df.columns),'CBM NAME',pd.NA)
df.insert(len(df.columns),'DIRECT RH',pd.NA)

df.columns
f1 = df[(df['RM DESIG 1']=='TL') & (df['RM DESIG 2']=='ABM ME') & (df['RM DESIG 3']=='CBM ME')]
f1_index = f1.index
df.loc[f1_index,['TL CODE','TL NAME','ABM CODE','ABM NAME','CBM CODE','CBM NAME']] = df.loc[f1_index,['RM CODE 1','RM NAME 1','RM CODE 2','RM NAME 2','RM CODE 3','RM NAME 3']].values
df.loc[f1_index,'DIRECT RH'] = '-'

#  FIlTER FOR 1 = TL AND 2 =ABM
f2 = df[(df.get('RM DESIG 1')=='TL') & (df.get('RM DESIG 2') =='ABM ME') & (df.get('TL CODE').isna())]
f2_index = f2.index
df.loc[f2_index,['TL CODE','TL NAME','ABM CODE','ABM NAME']]=df.loc[f2_index,['RM CODE 1','RM NAME 1','RM CODE 2','RM NAME 2']].values
df.loc[f2_index,['CBM CODE','CBM NAME','DIRECT RH']] = '-'

# FILTER FOR 1 = TL AND 2 = CBM
f3 = df[(df.get('RM DESIG 1') == 'TL') & (df.get('RM DESIG 2')=='CBM ME') & (df.get('TL CODE').isna())]
f3_index = f3.index
df.loc[f3_index,['TL CODE','TL NAME','CBM CODE','CBM NAME']] = df.loc[f3_index,['RM CODE 1','RM NAME 1','RM CODE 2','RM NAME 2']].values
df.loc[f3_index,['ABM CODE','ABM NAME','DIRECT RH']] = '-'

# 1 = TL AND 2 = TL
f4 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1') =='TL') & (df.get('RM DESIG 2')=='TL')]
f4_index  = f4.index
df.loc[f4_index,['TL CODE','TL NAME']] = df.loc[f4_index,['RM CODE 1','RM NAME 1']].values
df.loc[f4_index,['ABM CODE','ABM NAME','CBM CODE','CBM NAME','DIRECT RH']]  ='-'

# 1 = TL AND 3 = TL
f5 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1') =='TL') & (df.get('RM DESIG 3')=='TL')]
f5_index = f5.index
df.loc[f5_index,['TL CODE','TL NAME']] = df.loc[f5_index,['RM CODE 1','RM NAME 1']].values
df.loc[f5_index,['ABM CODE','ABM NAME','CBM CODE','CBM NAME','DIRECT RH']]  ='-'

# 1 = TL AND 3 = ABM
f6 = df[(df.get('RM DESIG 1')=='TL') & (df.get('RM DESIG 3')=='ABM ME') & (df.get('TL CODE').isna())]
f6_index= f6.index
df.loc[f6_index,['TL CODE','TL NAME','ABM CODE','ABM NAME']] = df.loc[f6_index,['RM CODE 1','RM NAME 1','RM CODE 3','RM NAME 3']].values
df.loc[f6_index,['CBM CODE','CBM NAME','DIRECT RH']] ='-'

# # 1 = TL AND 3 = CBM
f7 = df[(df.get('RM DESIG 1')=='TL') & (df.get('RM DESIG 3')=='CBM ME') & (df.get('TL CODE').isna())]
f7_index = f7.index
df.loc[f7_index,['TL CODE','TL NAME','CBM CODE','CBM NAME']] = df.loc[f7_index,['RM CODE 1','RM NAME 1','RM CODE 3','RM NAME 3']].values
df.loc[f7_index,['ABM CODE','ABM NAME','DIRECT RH']]='-'

f8 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='TL') & (df.get('RM DESIG 2')=='-') & (df.get('RM DESIG 3') =='-')]
f8_index = f8.index
df.loc[f8_index,['TL CODE','TL NAME']] = df.loc[f8_index,['RM CODE 1','RM NAME 1']].values
df.loc[f8_index,['ABM CODE','ABM NAME','CBM CODE','CBM NAME','DIRECT RH']] = '-'

f9 = df[(df.get('TL CODE').isna())& (df.get('RM DESIG 1')=='ABM ME') & (df.get('RM DESIG 2')=='CBM ME')]
f9_index = f9.index
df.loc[f9_index,['ABM CODE','ABM NAME','CBM CODE','CBM NAME']] = df.loc[f9_index,['RM CODE 1','RM NAME 1','RM CODE 2','RM NAME 2']].values
df.loc[f9_index,['TL CODE','TL NAME','DIRECT RH']]='-'

f10 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='ABM ME') & (df.get('RM DESIG 3')=='CBM ME')] 
f10_index = f10.index
df.loc[f10_index,['ABM CODE','ABM NAME','CBM CODE','CBM NAME']] = df.loc[f10_index,['RM CODE 1','RM NAME 1','RM CODE 3','RM NAME 3']].values
df.loc[f10_index,['TL CODE','TL NAME','DIRECT RH']] ='-'

f11 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='ABM ME') & (df.get('RM DESIG 2')=='ABM ME')]
f11_index = f11.index
df.loc[f11_index,['ABM CODE','ABM NAME']] = df.loc[f11_index,['RM CODE 1','RM NAME 1']].values
df.loc[f11_index,['TL CODE','TL NAME','CBM CODE','CBM NAME','DIRECT RH']]  = '-'

f12 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='ABM ME') & df.get('RM DESIG 3')=='ABM ME']
f12_index = f12.index
df.loc[f12_index,['ABM CODE','ABM NAME']] = df.loc[f12_index,['RM CODE 1','RM NAME 1']].values
df.loc[f12_index,['TL CODE','TL NAME','CBM CODE','CBM NAME','DIRECT RH']] = '-'

f13 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1') =='ABM ME') & (df.get('RM DESIG 2')=='-') & (df.get('RM DESIG 3')=='-')]
f13_index = f13.index
df.loc[f13_index,['ABM CODE','ABM NAME']] = df.loc[f13_index,['RM CODE 1','RM NAME 1']].values
df.loc[f13_index,['TL CODE','TL NAME','CBM CODE','CBM NAME','DIRECT RH']] = '-'

f14 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='CBM ME') & (df.get('RM DESIG 2')=='CBM ME')]
f14_index = f14.index
df.loc[f14_index,['CBM CODE','CBM NAME']]= df.loc[f14_index,['RM CODE 1','RM NAME 1']].values
df.loc[f14_index,['TL CODE','TL NAME','ABM CODE','ABM NAME','DIRECT RH']] = '-'

f15 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='CBM ME') & (df.get('RM DESIG 3')=='CBM ME')]
f15_index = f15.index
df.loc[f15_index,['CBM CODE','CBM NAME']] = df.loc[f15_index,['RM CODE 1','RM NAME 1']].values
df.loc[f15_index,['TL CODE','TL NAME','ABM CODE','ABM NAME','DIRECT RH']] = '-'

f16 = df[df.get('TL CODE').isna() & (df.get('RM DESIG 1')=='CBM ME') & (df.get('RM DESIG 2') =='-') & (df.get('RM DESIG 3')=='-')]
f16_index = f16.index
df.loc[f16_index,['CBM CODE','CBM NAME']] = df.loc[f16_index,['RM CODE 1','RM NAME 1']].values
df.loc[f16_index,['TL CODE','TL NAME','ABM CODE','ABM NAME','DIRECT RH']] = '-'


# from here the second is starting

f17 = df[(df.get('RM DESIG 2')=='TL') & (df.get('RM DESIG 3')=='ABM ME') & (df.get('TL CODE').isna())]
f17_index = f17.index
df.loc[f17_index,['TL CODE','TL NAME','ABM CODE','ABM NAME']] = df.loc[f17_index,['RM CODE 2','RM NAME 2','RM CODE 3','RM NAME 3']].values
df.loc[f17_index,['CBM CODE','CBM NAME','DIRECT RH']] = '-'

f22 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 2') == 'TL') & (df.get('RM DESIG 3')=='CBM ME')]
f22_index = f22.index
df.loc[f22_index,['TL CODE','TL NAME','CBM CODE','CBM NAME']] = df.loc[f22_index,['RM CODE 2','RM NAME 2','RM CODE 3','RM NAME 3']].values
df.loc[f22_index,['ABM CODE','ABM NAME','DIRECT RH']] = '-'

f18 = df[(df.get('RM DESIG 2')=='TL') & (df.get('TL CODE').isna()) & (df.get('RM DESIG 3')=='TL')]
f18_index = f18.index
df.loc[f18_index,['TL CODE','TL NAME']] = df.loc[f18_index,['RM CODE 2','RM NAME 2']].values
df.loc[f18_index,['ABM CODE','ABM NAME','CBM CODE','CBM NAME','DIRECT RH']] = '-'

f19 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 2')=='TL') & (df.get('RM DESIG 3') == '-')]
f19_index = f19.index
df.loc[f19_index,['TL CODE','TL NAME']] = df.loc[f19_index,['RM CODE 2','RM NAME 2']].values
df.loc[f19_index,['ABM CODE','ABM NAME','CBM CODE','CBM NAME','DIRECT RH']] = '-'

e1 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 2') == 'ABM ME') & (df.get('RM DESIG 3')=='CBM ME')]
e1_index = e1.index
df.loc[e1_index,['ABM CODE','ABM NAME','CBM CODE','CBM NAME']] = df.loc[e1_index,['RM CODE 2','RM NAME 2','RM CODE 3','RM NAME 3']].values
df.loc[e1_index,['TL CODE','TL NAME','DIRECT RH']] = '-'

e2 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 2')== 'ABM ME') & (df.get('RM DESIG 3') == 'ABM ME')]
e2_index = e2.index
df.loc[e2_index,['ABM CODE','ABM NAME']] = df.loc[e2_index,['RM CODE 2','RM NAME 2']].values
df.loc[e2_index,['TL CODE','TL NAME','CBM CODE','CBM NAME','DIRECT RH']] = '-'

e3 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 2')=='ABM ME') &  (df.get('RM DESIG 3')=='-')]
e3_index = e3.index
df.loc[e3_index,['ABM CODE','ABM NAME']]  = df.loc[e3_index,['RM CODE 2','RM NAME 2']].values
df.loc[e3_index,['TL CODE','TL NAME','CBM CODE','CBM NAME','DIRECT RH']] = '-'

e4 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 2')=='CBM ME') & (df.get('RM DESIG 3') == '-')]
e4_index = e4.index
df.loc[e4_index,['CBM CODE','CBM NAME']] = df.loc[e4_index,['RM CODE 2','RM NAME 2']].values
df.loc[e4_index,['TL CODE','TL NAME','ABM CODE','ABM NAME','DIRECT RH']] = '-'

e5 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 2')=='CBM ME') & (df.get('RM DESIG 3')=='CBM ME')]
e5_index = e5.index
df.loc[e5_index,['CBM CODE',"CBM NAME"]] = df.loc[e5_index,['RM CODE 2','RM NAME 2']].values
df.loc[e5_index,['TL CODE','TL NAME','ABM CODE','ABM NAME','DIRECT RH']] = '-'

#  last reporting manager now
c1 = df[(df.get('RM DESIG 3')=='ABM ME') & (df.get('TL CODE').isna())]
c1_index = c1.index
df.loc[c1_index,['ABM CODE','ABM NAME']] = df.loc[c1_index,['RM CODE 3','RM NAME 3']].values
df.loc[c1_index,['TL CODE','TL NAME','CBM CODE','CBM NAME','DIRECT RH']]='-'

c2 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 3')=='CBM ME')]
c2_index = c2.index
df.loc[c2_index,['CBM CODE','CBM NAME']] = df.loc[c2_index,['RM CODE 3','RM NAME 3']].values
df.loc[c2_index,['TL CODE','TL NAME','ABM CODE','ABM NAME','DIRECT RH']] = '-'


df.columns

r1 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='-') & (df.get('RM DESIG 2')=='-') & (df.get('RM DESIG 3')=='-') & (df.get('Role').isin({'BSM ME','SO ME'}))]
r1_index = r1.index
df.loc[r1_index,'DIRECT RH'] = df.loc[r1_index,'ME Region 2'].values
df.loc[r1_index,['TL CODE','TL NAME','ABM CODE','ABM NAME','CBM CODE','CBM NAME']] ='-'


r2 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='-') & (df.get('RM DESIG 2')=='-') & (df.get('RM DESIG 3')=='-') & (df.get('Role')=='TL')]
r2_index = r2.index
df.loc[r2_index,['TL CODE','TL NAME']] = df.loc[r2_index,['Employee_Code','Employee_Name']].values
df.loc[r2_index,['ABM CODE','ABM NAME','CBM CODE','CBM NAME','DIRECT RH']]='-'


r3 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='-') & (df.get('RM DESIG 2')=='-') & (df.get('RM DESIG 3')=='-') & (df.get('Role')=='ABM ME')]
r3_index = r3.index
df.loc[r3_index,['ABM CODE','ABM NAME']] = df.loc[r3_index,['Employee_Code','Employee_Name']].values
df.loc[r3_index,['TL CODE','TL NAME','CBM CODE','CBM NAME','DIRECT RH']] = '-'


r4 = df[(df.get('TL CODE').isna()) & (df.get('RM DESIG 1')=='-') & (df.get('RM DESIG 2')=='-') & (df.get('RM DESIG 3')=='-') & (df.get('Role')=='CBM ME')]
r4_index = r4.index
df.loc[r4_index,['CBM CODE','CBM NAME']] = df.loc[r4_index,['Employee_Code','Employee_Name']].values
df.loc[r4_index,['TL CODE','TL NAME','ABM CODE','ABM NAME','DIRECT RH']] = '-'




t1 = df[(df.get('TL CODE') != "-") & (df.get('ABM CODE')=='-')]
t1_index = t1.index
df.loc[t1_index,['ABM CODE','ABM NAME']]  = df.loc[t1_index,['TL CODE','TL NAME']].values
df.loc[t1_index,['TL CODE','TL NAME']] = '-'

# CAPITALIZING THE NAMES
df['Employee_Name'] = df['Employee_Name'].str.title()
df['ABM NAME'] = df['ABM NAME'].str.title()
df['TL NAME'] = df['TL NAME'].str.title()
df['CBM NAME'] = df['CBM NAME'].str.title()
df['RM NAME 1'] = df['RM NAME 1'].str.title()
df['RM NAME 2'] = df['RM NAME 2'].str.title()
df['RM NAME 3'] = df['RM NAME 3'].str.title()



df.columns

save_path = r"C:\Users\ET0001301\Pictures\Data\Manpower"
os.makedirs(save_path,exist_ok=True)
file_name = 'ABM_CBM From Code.xlsx'
final_path = os.path.join(save_path,file_name)
# df.to_excel(final_path)
with pd.ExcelWriter(final_path,engine='openpyxl') as writer:
    df.to_excel(writer,sheet_name='Data',index = False)
    wrong_emplyee.to_excel(writer,sheet_name='Wrong Employees',index=False)
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment
wb = load_workbook(final_path)
ws1 = wb['Data']

color = PatternFill(start_color='A9CCE3',fill_type='solid')
align = Alignment(horizontal='center',vertical='center')
bold = Font(bold= True)
for row in ws1.iter_rows():
    for cell in row:
        if cell.value:
            cell.alignment = align
for row in ws1[1]:
    row.fill = color
    row.font = bold
    
wb.save(final_path)
wb.close()
