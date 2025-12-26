import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import Font,Side,Border,PatternFill,Alignment
import win32com.client as win32
import shutil

# getting direclty data from the oulook
outlook = win32.Dispatch('Outlook.Application').GetNamespace('MAPI')
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items
messages.sort("[ReceivedTime]",True)

mail_file_save_path = r"C:\Users\et0001301\Desktop\Python\OUTLOOK DATA"

subject_date_time = pd.Timestamp.now().strftime('%d%b%Y')

os.makedirs(mail_file_save_path,exist_ok = True)
mail_file_name = f'Declined_data-{subject_date_time}.xlsx'

outlook_path_to_save_file = os.path.join(mail_file_save_path,mail_file_name)


for mail in messages:
    subject = mail.subject
    if  'ME-MIS-' in subject and subject_date_time in subject:
        if mail.Attachments.count>=2:
            second_attachment = mail.Attachments.Item(2)
            second_attachment.SaveAsFile(outlook_path_to_save_file)
        source_subject = mail.Subject
        source_sender = mail.SenderEmailAddress
        source_received_time = mail.ReceivedTime.strftime('%d-%b-%Y %H:%M:%S')   
        mail_info = {
                    "subject":mail.subject,
                    "sender":mail.sender,
                    'Received_time':mail.ReceivedTime
                    }       
        for k,v in mail_info.items():
            print(f'{k}:{v}')
        break            


main_data_path = r"C:\Users\et0001301\Desktop\Python\MAIN DATA\Decline Data"
os.makedirs(main_data_path,exist_ok = True)
final_path = os.path.join(main_data_path,mail_file_name)

shutil.copy(outlook_path_to_save_file,main_data_path)

df = pd.read_excel(final_path,sheet_name = 'out_file')

df['DOCUMENTRECEIVEDATCPADATE']= pd.to_datetime(df['DOCUMENTRECEIVEDATCPADATE'])

current_month_for_filter = pd.Timestamp.now().month
current_year_for_filter = pd.Timestamp.now().year
df = df[(df['MIS_STATUS']=='Declined') & (df['DOCUMENTRECEIVEDATCPADATE'].dt.month == current_month_for_filter) & (df['DOCUMENTRECEIVEDATCPADATE'].dt.year == current_year_for_filter)]

df.insert(13,'Cibil_bucketing',pd.NA)

df['Cibil_bucketing'] = np.select([(df['APPLICANTCIBILSCORE']==-1),(df['APPLICANTCIBILSCORE']>=700)],
                                  ['-1','>=700'],
                                  default = ""
                                  )
df.insert(28,'Logins',pd.NA)

df['Logins'] = df['DOCUMENTRECEIVEDATCPADATE'].apply(lambda x:pd.Timestamp.now().strftime('%b') if ((x.month ==current_month_for_filter) and (x.year == current_year_for_filter)) else "")

df['LAST_MASTER_REMARK'] = df['LAST_MASTER_REMARK'].str.strip()

df['LAST_MASTER_REMARK'] = df['LAST_MASTER_REMARK'].fillna('No Remark')

east_data = df[(df['ZONE'].str.lower()=='east') & (df['Cibil_bucketing'].isin({'-1','>=700'})) & (df['MIS_STATUS']=='Declined') & (df['Logins']==pd.Timestamp.now().strftime('%b'))]
west_data = df[(df['ZONE'].str.lower()=='west') & (df['Cibil_bucketing'].isin({'-1','>=700'})) & (df['MIS_STATUS']=='Declined') & (df['Logins']==pd.Timestamp.now().strftime('%b'))]
north_data = df[(df['ZONE'].str.lower()=='north') & (df['Cibil_bucketing'].isin({'-1','>=700'})) & (df['MIS_STATUS']=='Declined') & (df['Logins']==pd.Timestamp.now().strftime('%b'))]
south_data = df[(df['ZONE'].str.lower()=='south') & (df['Cibil_bucketing'].isin({'-1','>=700'})) & (df['MIS_STATUS']=='Declined') & (df['Logins']==pd.Timestamp.now().strftime('%b'))]

file_name_time = pd.Timestamp.now().strftime('%d-%b-%Y')

main_folder_path  = r"C:\Users\et0001301\Desktop\Python\MAIN DATA\Decline Data"
# main_folder_path = os.path.join(test_folder_path,file_name_time)
os.makedirs(main_folder_path,exist_ok = True)
east_file_name = 'East_Declined_data.xlsx'
west_file_name = 'West_Declined_data.xlsx'
north_file_name = 'North_Declined_data.xlsx'
south_file_name = 'South_Declined_data.xlsx'


east_final_path = os.path.join(main_folder_path,east_file_name)
west_final_path = os.path.join(main_folder_path,west_file_name)
north_final_path = os.path.join(main_folder_path,north_file_name)
south_final_path = os.path.join(main_folder_path,south_file_name)


east_data.to_excel(east_final_path,sheet_name = 'DATA',index = False)
west_data.to_excel(west_final_path,sheet_name = 'DATA',index = False)
north_data.to_excel(north_final_path,sheet_name = 'DATA',index = False)
south_data.to_excel(south_final_path,sheet_name = 'DATA',index = False)

def create_pivot(data):
    pivot = pd.pivot_table(data,
                           index = 'REGION',
                           columns = 'LAST_MASTER_REMARK',
                           values = 'REFERENCEID',
                           aggfunc= 'count',
                           margins = True,
                           margins_name = 'Grand Total',
                           ).reset_index() 
    pivot = pivot.fillna('-')
    return pivot

east_pivot = create_pivot(east_data)
west_pivot = create_pivot(west_data)
north_pivot = create_pivot(north_data)
south_pivot = create_pivot(south_data)

with pd.ExcelWriter(east_final_path,engine = 'openpyxl',mode = 'a',if_sheet_exists = 'overlay') as writer:
    east_pivot.to_excel(writer,sheet_name = 'Summary',index = False)
    

with pd.ExcelWriter(west_final_path,engine = 'openpyxl',mode = 'a',if_sheet_exists = 'overlay') as writer:
    west_pivot.to_excel(writer,sheet_name = 'Summary',index = False)
    

with pd.ExcelWriter(north_final_path,engine = 'openpyxl',mode = 'a',if_sheet_exists = 'overlay') as writer:
    north_pivot.to_excel(writer,sheet_name = 'Summary',index = False)


with pd.ExcelWriter(south_final_path,engine = 'openpyxl',mode = 'a',if_sheet_exists = 'overlay') as writer:
    south_pivot.to_excel(writer,sheet_name = 'Summary',index = False)  
    
              
# Now formatting the each sheet
side = Side(style = 'thin')
border  = Border(left = side,
                 right = side,
                 top = side,
                 bottom = side)
alignment = Alignment(horizontal = 'center',vertical = 'center')
header_fill = PatternFill(start_color = '87CEEB',fill_type = 'solid')
total_fill = PatternFill(start_color = '87CEEB',fill_type ='solid')
header_font = Font(bold = True,name = 'Aptos Narrow')
value_font = Font(name ='Aptos Narrow')

def format(path):
    wb = load_workbook(path)  
    ws1 = wb['DATA']
    ws2 = wb['Summary']
    
    def data_column_width(ws1):
        for col in ws1.iter_cols():
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length,len(str(cell.value)))
            ws1.column_dimensions[col_letter].width = max_length + 3
    def summary_column_width(ws2):
        for col in ws2.iter_cols():
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length,len(str(cell.value)))
            ws2.column_dimensions[col_letter].width = max_length + 3
    data_column_width(ws1)
    summary_column_width(ws2)
    
    def data_sheet_format(ws1): 
        for cell in ws1[1]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment
        for col in ws1.iter_cols():
            for cell in col:
                cell.border = border
                cell.alignment = alignment
                cell.font = value_font
    
    data_sheet_format(ws1)
                        
    def summary_sheet_format(ws2):
        for col in ws2.iter_cols(min_row = 1,max_row = 1,min_col = 1,max_col = ws2.max_column):
            for cell in col:
                if cell.value is not None:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = alignment
        for col in ws2.iter_cols(min_row = 2,max_row = ws2.max_row,min_col = 1,max_col = ws2.max_column):
            for cell in col:
                cell.border = border
                cell.alignment = alignment
                cell.font = value_font
        for row in ws2.iter_rows(min_row = ws2.max_row,max_row = ws2.max_row,min_col = 1,max_col = ws2.max_column):
            if row[0].value == 'Grand Total':
                for cell in row:
                    cell.fill = total_fill
                    cell.alignment = alignment
                    cell.font = header_font  
    summary_sheet_format(ws2)
    wb.save(path)
    wb.close()                                               

format(east_final_path)    
format(west_final_path)    
format(north_final_path)    
format(south_final_path)    


current_time_for_mail = pd.Timestamp.now().strftime('%d%b%Y')

send_outlook = win32.Dispatch('Outlook.Application')
mail = send_outlook.CreateItem(0)

mail.to = 'ravi.gupta@sbfc.com'
mail.subject = f'DECLINE DATA AS ON {current_time_for_mail}'
mail.HTMLBody = """
<html>
<body>
<p> Hello Ravi,</p>

<p> Please find the Declined  data </p>

<p> Regards,</p>
<p> Ravi Gupta</p>


</body>
</html> 

"""

files = [east_final_path,west_final_path,north_final_path,south_final_path]
for att in files:
    mail.Attachments.Add(att) 
mail.Send()
print('ALL IS DONE')        
