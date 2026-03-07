import win32com.client as win
import pythoncom
pythoncom.CoInitialize()
outlook = win.Dispatch('Outlook.Application').GetNameSpace('MAPI')
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items
messages.Sort('[ReceivedTime]',True)

from datetime import timedelta
date_for_outlook = pd.Timestamp.now()
yest_date = date_for_outlook-timedelta(days=1)
yest_date_format = yest_date.strftime('%d %b %y')
import os
outlook_save_path = r"C:\Users\ET0001301\Pictures\Data\Closure on Day\Outlook"
os.makedirs(outlook_save_path,exist_ok=True)
file_name = f'Daily closure {yest_date_format}.xlsx'
outlook_save_path2 = os.path.join(outlook_save_path,file_name)

Yaha par date change karna hai maine manually 5 ka date dala hai

for mail in messages:
    sub = mail.subject
    if 'ME_DAILY_ClOSURE_MIS' in sub and yest_date_format in sub:
        if mail.Attachments.count>=2:
            second_file = mail.Attachments.Item(2)
            second_file.saveAsFile(outlook_save_path2)
        mail_info = {'sender':mail.sender,
                     'Sub':mail.Subject,
                     'Received Time':mail.ReceivedTime
                     }
        
        for k,v in mail_info.items():
            print(k,v)
        break    
import shutil
shutil_path = r"C:\Users\ET0001301\Pictures\Data\Shutil Data\Closure Shutil"
os.makedirs(shutil_path,exist_ok=True)
shutil_file = f'Daily closure {yest_date_format}.xlsx'
shutil.copy(outlook_save_path2,shutil_path)
shutil_final = os.path.join(shutil_path,shutil_file)
import pandas as pd
# path = r"C:\Users\ET0001301\Pictures\DailyClosurePivot.xlsm"
df = pd.read_excel(shutil_final,sheet_name='Sheet1')
pos_path = r"C:\Users\ET0001301\Pictures\Data\Closure on Day\1.LAP_PORTFOLIO_FEB26.xlsb"
pos = pd.read_excel(pos_path,sheet_name='Sheet1')
df= df.merge(pos[['PROSPECTCODE','POS IN Cr']],on = 'PROSPECTCODE', how='left')
df['POS IN Cr'].round(2)
df = df[df['ProductCheck'] != 'Rest']
df = df[df['REASON CODE'].isin(['CANCELLATION','NORMAL FORECLOSURE'])]
region = ['Assam','WB','Bihar','Delhi','Haryana','PCH','Rajasthan','Uttarakhand','UP East','UP West','Varanasi','Bangalore','ROK','Telangana 1',
'Telangana 2',
'AP',
'TN',
'Mumbai',
'ROM',
'ROM 2',
'Gujarat',
'MP'
]
region
type(region)
region2 = pd.Series(region,name = 'Region')
region3 = region2.to_frame()
region3
type(region3)
df
grp1 = df.groupby(['Region2','REASON CODE'])['POS IN Cr'].agg('sum').unstack()
grp1.columns.Name = None
grp1
region4 = region3.merge(grp1,left_on = 'Region',right_on='Region2',how ='left')
region4.fillna(0)
region4['Grand Total'] = region4.sum(axis = 1,numeric_only=True)
region4 = region4.fillna(0).round(2)
target = {
    'Assam':0.82,'WB':0.04,'Bihar':2.32,
    'Delhi':3.15,'Haryana':0.77,'PCH':0.21,'Rajasthan':0.55,'Uttarakhand':1.18,
    'UP East':1.21,'UP West':1.73,'Varanasi':0.77,
    'Bangalore':3.19,'ROK':0.62,
    'Telangana 1':1.11,'Telangana 2':2.16,
    'AP':1.49,'TN':0.98,
    'Mumbai':0.59,'ROM':0.29,'ROM 2':0.52,
    'Gujarat':0.47,'MP':1.98
}
target_df = pd.DataFrame(list(target.items()),columns = ['Region','Max Closure'])
region4
region4 = region4.merge(target_df,on = 'Region',how = 'left')
region4 = region4.rename(columns = {'NORMAL FORECLOSURE':'Closure'})
region4
east_total = pd.DataFrame({'Region':['East Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['Assam','WB','Bihar']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['Assam','WB','Bihar']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['Assam','WB','Bihar']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['Assam','WB','Bihar']),'Max Closure'].sum()]
                           })


pos1 = region4.loc[region4['Region']=='Bihar'].index.max()+1
region4 = pd.concat([region4.iloc[:pos1],east_total,region4.iloc[pos1:]],ignore_index=True)
region4

delhi_total = pd.DataFrame({'Region':['Delhi Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['Delhi','Haryana','PCH','Rajasthan','Uttarakhand']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['Delhi','Haryana','PCH','Rajasthan','Uttarakhand']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['Delhi','Haryana','PCH','Rajasthan','Uttarakhand']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['Delhi','Haryana','PCH','Rajasthan','Uttarakhand']),'Max Closure'].sum()]
                           })
  
pos2 = region4.loc[region4['Region']=='Uttarakhand'].index.max()+1
region4 = pd.concat([region4.iloc[:pos2],delhi_total,region4.iloc[pos2:]],ignore_index=True)
region4

up_total = pd.DataFrame({'Region':['UP Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['UP East','UP West','Varanasi']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['UP East','UP West','Varanasi']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['UP East','UP West','Varanasi']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['UP East','UP West','Varanasi']),'Max Closure'].sum()]
                           })
  
pos3 = region4.loc[region4['Region']=='Varanasi'].index.max()+1
region4 = pd.concat([region4.iloc[:pos3],up_total,region4.iloc[pos3:]],ignore_index = True) 
region4

north_total = pd.DataFrame({'Region':['North Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['UP Total','Delhi Total']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['UP Total','Delhi Total']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['UP Total','Delhi Total']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['UP Total','Delhi Total']),'Max Closure'].sum()]
                           })
  
pos4 = region4.loc[region4['Region']=='UP Total'].index.max()+1
region4 = pd.concat([region4.iloc[:pos4],north_total,region4.iloc[pos4:]],ignore_index=True)
region4

karnataka_total = pd.DataFrame({'Region':['Karnataka Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['Bangalore','ROK']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['Bangalore','ROK']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['Bangalore','ROK']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['Bangalore','ROK']),'Max Closure'].sum()]
                           })
  
pos5 = region4.loc[region4['Region']=='ROK'].index.max()+1
region4 = pd.concat([region4.iloc[:pos5],karnataka_total,region4.iloc[pos5:]],ignore_index=True)
region4

telangana_total = pd.DataFrame({'Region':['Telangana Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['Telangana 1','Telangana 2']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['Telangana 1','Telangana 2']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['Telangana 1','Telangana 2']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['Telangana 1','Telangana 2']),'Max Closure'].sum()]
                           })
  
pos6 = region4.loc[region4['Region']=='Telangana 2'].index.max()+1
region4 = pd.concat([region4.iloc[:pos6],telangana_total,region4.iloc[pos6:]],ignore_index=True)

south_total = pd.DataFrame({'Region':['South Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['Karnataka Total','Telangana Total','AP','TN']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['Karnataka Total','Telangana Total','AP','TN']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['Karnataka Total','Telangana Total','AP','TN']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['Karnataka Total','Telangana Total','AP','TN']),'Max Closure'].sum()]
                           })

pos7 = region4.loc[region4['Region']=='TN'].index.max()+1
region4 = pd.concat([region4.iloc[:pos7],south_total,region4.iloc[pos7:]],ignore_index=True)

mh_total = pd.DataFrame({'Region':['Maharashtra Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['Mumbai','ROM','ROM 2']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['Mumbai','ROM','ROM 2']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['Mumbai','ROM','ROM 2']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['Mumbai','ROM','ROM 2']),'Max Closure'].sum()]
                           })

pos8 = region4.loc[region4['Region']=='ROM 2'].index.max()+1
region4 = pd.concat([region4.iloc[:pos8],mh_total,region4.iloc[pos8:]],ignore_index=True)


west_total = pd.DataFrame({'Region':['West Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['Maharashtra Total','Gujarat','MP']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['Maharashtra Total','Gujarat','MP']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['Maharashtra Total','Gujarat','MP']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['Maharashtra Total','Gujarat','MP']),'Max Closure'].sum()]
                           })
  
pos9 = region4.loc[region4['Region']=='MP'].index.max()+1
region4 = pd.concat([region4.iloc[:pos9],west_total,region4.iloc[pos9:]],ignore_index=True)

grand_total = pd.DataFrame({'Region':['Grand Total'],
                           'CANCELLATION':[region4.loc[region4['Region'].isin(['East Total','North Total','South Total','West Total']),'CANCELLATION'].sum()],
                           'Closure':[region4.loc[region4['Region'].isin(['East Total','North Total','South Total','West Total']),'Closure'].sum()],
                           'Grand Total':[region4.loc[region4['Region'].isin(['East Total','North Total','South Total','West Total']),'Grand Total'].sum()],
                           'Max Closure':[region4.loc[region4['Region'].isin(['East Total','North Total','South Total','West Total']),'Max Closure'].sum()]
                           })
  
pos10 = region4.loc[region4['Region']=='West Total'].index.max()+1
region4 = pd.concat([region4.iloc[:pos10],grand_total,region4.iloc[pos10:]],ignore_index=True)
region4 = region4.fillna(0)
region4['Max Closure %'] = (region4['Grand Total'].div(region4['Max Closure']).mul(100)).fillna(0).astype(int).astype(str)+'%'
region4
save_file_date = pd.Timestamp.now()
yest_date_for_save = (save_file_date-timedelta(days=1)).strftime('%d %b %y')
yest_date_for_save
save_path = r"C:\Users\ET0001301\Pictures\Data\Closure on Day"
main_file_name = f'Daily Closure {yest_date_for_save}.xlsx'
os.makedirs(save_path,exist_ok=True)
main_save_path = os.path.join(save_path,main_file_name)
with pd.ExcelWriter(main_save_path,engine='openpyxl') as writer:
    df.to_excel(writer,sheet_name='Data',index = False)
    region4.to_excel(writer,sheet_name = 'Summary',startrow = 1,index = False)
from openpyxl.styles import Font,Side,PatternFill,Alignment,Border
from openpyxl import load_workbook
align = Alignment(horizontal='center',vertical='center')
side = Side(style='thin')
border = Border(left = side,
                right = side,
                top = side,
                bottom = side)
color = PatternFill(start_color='A9CCE3',fill_type='solid')
bold = Font(bold = True)
wb = load_workbook(main_save_path)
ws = wb['Summary']

merge_rows = ['A1:A2','B1:B2','C1:C2','D1:D2','E1:E2','F1:F2']
ws['A1']='Region'
ws['B1']='CANCELLATION'
ws['C1']='Closure'
ws['D1'] = 'Grand Total'
ws['E1'] = 'Max Closure'
ws['F1'] = 'Max Closure %'
target_cols = ['CANCELLATION','Closure','Grand Total','Max Closure']
for cols in ws.iter_cols(min_row=1):
    header = cols[0].value
    if header in target_cols:
        for cells in cols[1:]:
            if isinstance(cells.value,(int,float)):
                cells.number_format = '0.00'

for cell in merge_rows:
    ws.merge_cells(cell)
for row in ws.iter_rows(min_row=1,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
    for cell in row:
        cell.alignment = align
        cell.border = border
            
for cols in ws.iter_cols(min_col=1,max_col=ws.max_column,min_row=1,max_row=ws.max_column):
    max_length = 0
    col_letter = cols[0].column_letter
    for cell in cols:
        if cell.value:
            max_length = max(max_length,len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length+1           
for row in ws[1]:
    row.font= bold
    row.fill = color
        
format = ['East Total','Delhi Total','UP Total','North Total','Karnataka Total','Telangana Total','South Total',
          'Maharashtra Total','West Total','Grand Total']
for cols in ws.iter_rows():
    for cell in cols:
        if cell.value in format: 
            for cell in cols:
                cell.fill = color
wb.save(main_save_path)
wb.close()
outlook1 = win.Dispatch('Outlook.Application')
mail= outlook1.CreateItem(0)
mail.to = 'ravi.gupta@sbfc.com'
mail.subject = f'Daily Closure {yest_date_for_save}'
mail.Attachments.Add(main_save_path)
mail.Display()
mail.Send()
