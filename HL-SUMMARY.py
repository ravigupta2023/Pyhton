import pandas as pd
import os
import time 
from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side,Alignment,PatternFill
from openpyxl.utils import get_column_letter
import win32com.client as win32
import shutil

print("THE PROGRAM HAS BEEN STARTED")

mail_start_time = pd.Timestamp.now().strftime('%H:%M:%S')
print(f'MAIL STARTED TIME IS {mail_start_time}')

outlook = win32.Dispatch('Outlook.Application').GetNameSpace('MAPI')
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items

messages.Sort("[ReceivedTime]",True)

subject_current_date = pd.Timestamp.now().strftime('%d%b%Y')

print('GETTING THE MAIL FROM THE OUTLOOK')

outlook_path = r"C:\Users\et0001301\Desktop\Python\OUTLOOK DATA"
os.makedirs(outlook_path,exist_ok = True)
outlook_file_name = f'HL-DATA-{subject_current_date}.xlsx'

outlook_file_path = os.path.join(outlook_path,outlook_file_name)

for mail in messages:
    subject = mail.Subject 
    if 'HL-MIS-' in subject and subject_current_date in subject:
        if mail.Attachments.count>=2:
            second_attachment = mail.Attachments.Item(2)
            second_attachment.SaveAsFile(outlook_file_path)
        mail_info = {
                    "Subject":mail.Subject,
                    'Sender':mail.sender,
                    'Received Time': mail.ReceivedTime
                    }
        for k,v in mail_info.items():
            print(f'{k}:{v}')
        break

time.sleep(2)
print('THE LATEST MAIL HAS BEEN FOUND ')
main_folder_path = r"C:\Users\et0001301\Desktop\Python\MAIN DATA"
os.makedirs(main_folder_path,exist_ok = True)
file_name = outlook_file_name

shutil.copy(outlook_file_path,main_folder_path)  
main_path = os.path.join(main_folder_path,file_name)

print('PATH COPIED INTO THE MAIN DATA FOLDER')

# Getting the path of the data
# data_path = r"C:\Users\et0001301\Documents\ME-MIS_HL_SUMMARY.xlsx"
df = pd.read_excel(main_path,sheet_name = 'out_file')

print('LOADING THE DATA INTO THE PANDAS DATAFRAME FOR CLEANING')
df = df[df['REGION'].str.lower()=='gujarat']
df['Tranch'] = df['SCHEMETYPE'].apply(lambda x:'YES' if x == 'HL BT Top up' else "NO")
df['MIS_STATUS'] = df.apply(lambda x:'WIP' if 
                                            ((x['CURRENTSTATUS'] in {'Hold','Queue'}) & (x['MIS_STATUS']=='Approved')) 
                                            else x['MIS_STATUS'],axis = 1
                                            )

df['login_date'] = pd.to_datetime(df['login_date'])
df['DOCUMENTRECEIVEDATCPADATE']=pd.to_datetime(df['DOCUMENTRECEIVEDATCPADATE'])

df.insert(12,'Decision_month',pd.NA)
df.insert(28,'Logins',pd.NA)

current_date = pd.Timestamp.now()
current_month_name = current_date.strftime('%b')
current_month = current_date.month
current_year = current_date.year

df['Decision_month'] = df['login_date'].apply(lambda x:current_month_name if ((x.month == current_month) & (x.year == current_year)) else "")
df['Logins'] = df['DOCUMENTRECEIVEDATCPADATE'].apply(lambda x:current_month_name if ((x.month ==current_month ) & (x.year == current_year)) else "")

file_date_month_name = pd.Timestamp.now().strftime('%d-%b-%Y')

# save_folder = r"C:\Users\et0001301\Desktop\Python\MAIN DATA"
# os.makedirs(save_folder,exist_ok = True)
# file_name  = f'HL-Outlook_data-{file_date_month_name}.xlsx'

# main_path = os.path.join(save_folder,file_name)

df.to_excel(main_path,sheet_name = "DATA" , index = False)

# creating pivot
print('DATA SHEET HAS BEEN CREATE AND CLEANED')

filter_for_logins = df[(df['Tranch']=='NO') & (df['Logins'] == current_month_name)]
filter_for_app_subapp_wip = df[(df['Tranch']=='NO') & (df['Decision_month']==current_month_name) & (df['MIS_STATUS'].isin({'Approved','Subjective Approval','WIP'}))]
filter_for_disb_dec = df[(df['Decision_month']==current_month_name) & (df['MIS_STATUS'].isin({'Disbursed','Declined'}))]

print('CREATING PIVOT')

pivot_for_logins = pd.pivot_table(filter_for_logins,
                                  index = 'BRANCH',
                                  values = 'IN_Cr',
                                  aggfunc = ['count','sum'],
                                  margins = True,
                                  margins_name = 'Grand Total',
                                  fill_value = 0
                                  )
pivot_for_first_three_mis_status = pd.pivot_table(filter_for_app_subapp_wip,
                                                  index = 'BRANCH',
                                                  columns = 'MIS_STATUS',
                                                  values = 'IN_Cr',
                                                  aggfunc = ['count','sum'],
                                                  margins = True,
                                                  margins_name = 'Grand Total',
                                                  fill_value = 0
                                                  )
pivot_for_disb_declined = pd.pivot_table(
                                        filter_for_disb_dec,
                                        index = 'BRANCH',
                                        columns = 'MIS_STATUS',
                                        values = 'IN_Cr',
                                        aggfunc = ['count','sum'],
                                        margins = True,
                                        margins_name = 'Grand Total',
                                        fill_value = 0
                                         )
with pd.ExcelWriter (main_path,engine = 'openpyxl',mode = 'a',if_sheet_exists = 'overlay') as writer:
    table = [('LOGINS_SUMMARY',pivot_for_logins),
             ('APP_SUBAPP_WIP_SUMMARY',pivot_for_first_three_mis_status),
             ('DISBURSAL AND DECLINED SUMMARY',pivot_for_disb_declined)
             ]
    
    
    row = 0
    col = 0
    for title,pivot in table:
        ws = writer.book.create_sheet('PIVOTS') if 'PIVOTS' not in writer.book.sheetnames else writer.book['PIVOTS']
        ws.cell(row =row+1,column = col+1 ).value = title
        pivot.to_excel(writer,sheet_name = 'PIVOTS',startrow = row +1,startcol = col)
        col += pivot.shape[1]+2
        
time.sleep(2)
print('STORING THE PIVOT IN THE SHEET')
# formatting the pivot and data sheet 
wb = load_workbook(main_path)
ws1 = wb['DATA']
ws2 = wb['PIVOTS']

# Defining the formatting variables font color like this 
print('LOADING THE DATA INTO THE OPENPYXL FOR FORMATTING')

header_color = PatternFill(start_color='87CEEB',fill_type = 'solid')
font = Font(bold = True)
side = Side(style = 'thin')
border = Border(
                left = side,
                right = side,
                top = side,
                bottom = side
                )
alignment = Alignment(horizontal = 'center',vertical = 'center')

for col in ws1.iter_cols():   # FITTING THE DATA SHEET COLUMN WIDTH 
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value is not None:
            max_length = max(max_length,len(str(cell.value)))
    ws1.column_dimensions[col_letter].width = max_length + 4        


for col in ws2.iter_cols():   # FITTING THE PIVOT SHEET COLUMN WIDTH 
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value is not None:
            max_length = max(max_length,len(str(cell.value)))
    ws2.column_dimensions[col_letter].width = max_length + 4    
    
for cell in ws1[1]:  # header color of data sheet
    if cell.value is not None:
        cell.fill = header_color
        
for row in ws1.iter_rows():  # formatting the data sheet value not header
    for cell in row:
        cell.border = border
        cell.alignment = alignment

for cell in ws2[1]:              # Highlighting the first row of pivot 
    if cell.value is not None:
        cell.fill = PatternFill(start_color = '87EE5B',fill_type = 'solid')
        cell.font = font

ws2.merge_cells("B2:B3")    # merging the two cell of the login Summary
ws2.merge_cells('C2:C3')

for row in ws2[2:3]:    # Highlighting the 2nd  and third row that is header of pivot
    for cell in row:
        if cell.value is not None:
            cell.fill = header_color
            
for row in ws2.iter_rows(min_row = 2, max_row = ws2.max_row,min_col = 1,max_col = ws2.max_column):   # formatting the values of the pivot sheet
    for cell in row:
        if cell.value is not None:
            cell.border = border
            cell.alignment = alignment
            
wb.save(main_path)
wb.close()

print('PIVOT FORMATTING DONE')
# creating the summary sheet
print('NOW CREATING THE SUMMARY SHEET')
summary_header = ['Branch','Logins','Approved','Subjective Approval','WIP','Disbursed','Declined','Total'] 

wb2 = load_workbook(main_path)
ws3 = wb2.create_sheet('SUMMARY') if 'SUMMARY' not in wb2.sheetnames else wb2['SUMMARY']

row1 = 1
col1 = 1 

for header in summary_header:
    if header == 'Branch':
        ws3.cell(row = row1,column = col1,value = header)
        ws3.merge_cells(start_row = row1,end_row = row1+1,start_column = col1,end_column = col1 )
        col1 +=1
    else:
        ws3.cell(row = row1,column = col1,value = header)
        ws3.merge_cells(start_row = row1,end_row = row1,start_column = col1,end_column = col1+1)
        col1 += 2
        
# sub headers that is count and sum

col2 = 2 
for _ in summary_header[1:]:
    ws3.cell(row = 2,column = col2,value = 'count')
    ws3.cell(row = 2,column = col2+1,value = 'sum')
    col2 +=2


# Branch name
branches_consolidated = ['Anand','Gurukul','Maninagar','Mavdi','Mehsana','Morbi','Palanpur','Rajkot','Silvassa','Surat','Vadodara','Vapi','Total']

row3 = 3
col3 = 1

for branch in branches_consolidated:  # Putting branch names
    ws3.cell(row = row3,column = col3,value = branch)
    row3 +=1
    
# fitting the branch header and values

Blue_Code = PatternFill(start_color = '95F5F7',fill_type = 'solid')
other_header_color = PatternFill(start_color = 'CEBECE',fill_type = 'solid')
Side1 = Side(style = 'thin')
summary_border = Border(
                        left = Side1,
                        right = Side1,
                        top = Side1,
                        bottom = Side1
                        )
summary_alignment = Alignment(horizontal = 'center',vertical = 'center')

for row in ws3.iter_rows():    # formatting the values of the summary sheet
    for cell in row:
        cell.border = summary_border
        cell.alignment = alignment
        
for row in ws3.iter_rows(min_row = 1,max_row = ws3.max_row,min_col = 1,max_col = 1):  # Blue color for branch name
       for cell in row:
           if cell.value == 'Branch':
               cell.fill= Blue_Code
           if cell.value == 'Total' :
               cell.fill = Blue_Code
               
for row in ws3.iter_rows(min_row = 1,max_row = 2,min_col = 2,max_col = ws3.max_column):   # formatting the remaining headers                        
       for cell in row:
           cell.fill = other_header_color
               
for col in ws3.iter_cols():                   # formatting the other row of the total header
    for idx,cell in enumerate(col,start = 1):
        if cell.value == 'Total':
            start_col = cell.column
            for c in range(start_col+1,ws3.max_column+1):
                cells = ws3.cell(row = 15,column = c)
                cells.fill = other_header_color
                
                
for col in ws3.iter_cols():  # AutoColumn width fitting
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value is not None:
            max_length = max(max_length,len(str(cell.value)))
            
    ws3.column_dimensions[col_letter].width = max_length + 3
    
    
print('SUMMARY SHEET CREATION AND FORMATTING IS DONE')
    
ws3['B3']= "=SUMIF(PIVOTS!$A:$A,A3,PIVOTS!$B:$B)"  # LOOPING VALUE FROM THE PIVOT SHHEET INTO THE SUMMARY SHEET
ws3['C3']= "=SUMIF(PIVOTS!$A:$A,A3,PIVOTS!$C:$C)"
ws3['D3']= "=SUMIF(PIVOTS!$E:$E,A3,PIVOTS!$F:$F)"
ws3['E3']= "=SUMIF(PIVOTS!$E:$E,A3,PIVOTS!$J:$J)"
ws3['F3']= "=SUMIF(PIVOTS!$E:$E,A3,PIVOTS!$G:$G)"
ws3['G3']= "=SUMIF(PIVOTS!$E:$E,A3,PIVOTS!$K:$K)"
ws3['H3']= "=SUMIF(PIVOTS!$E:$E,A3,PIVOTS!$H:$H)"
ws3['I3']= "=SUMIF(PIVOTS!$E:$E,A3,PIVOTS!$L:$L)"
ws3['J3']= "=SUMIF(PIVOTS!$O:$O,A3,PIVOTS!$Q:$Q)"
ws3['K3']= "=SUMIF(PIVOTS!$O:$O,A3,PIVOTS!$T:$T)"
ws3['L3']= "=SUMIF(PIVOTS!$O:$O,A3,PIVOTS!$P:$P)"
ws3['M3']= "=SUMIF(PIVOTS!$O:$O,A3,PIVOTS!$S:$S)"

ws3['N3']= "=SUM(D3,F3,J3)"  # SUM values on the total column
ws3['O3']= "=SUM(E3,G3,K3)"

for row in range(3,15):              #fillng the TOTAL VALUEs
    ws3[f"N{row}"] = f"=SUM(D{row},F{row},J{row})"
    ws3[f"O{row}"]= f"=SUM(E{row},G{row},K{row})"
    
for row in range(3,15):
    ws3[f"B{row}"] = ws3['B3'].value.replace('A3',f'A{row}')                 
    ws3[f"C{row}"] = ws3['C3'].value.replace('A3',f'A{row}')                 
    ws3[f"D{row}"] = ws3['D3'].value.replace('A3',f'A{row}')                 
    ws3[f"E{row}"] = ws3['E3'].value.replace('A3',f'A{row}')                 
    ws3[f"F{row}"] = ws3['F3'].value.replace('A3',f'A{row}')                 
    ws3[f"G{row}"] = ws3['G3'].value.replace('A3',f'A{row}')                 
    ws3[f"H{row}"] = ws3['H3'].value.replace('A3',f'A{row}')                 
    ws3[f"I{row}"] = ws3['I3'].value.replace('A3',f'A{row}')                 
    ws3[f"J{row}"] = ws3['J3'].value.replace('A3',f'A{row}')                 
    ws3[f"K{row}"] = ws3['K3'].value.replace('A3',f'A{row}')                 
    ws3[f"L{row}"] = ws3['L3'].value.replace('A3',f'A{row}')                 
    ws3[f"M{row}"] = ws3['M3'].value.replace('A3',f'A{row}')                 
        

# summing all the values
ws3['B15'] = "=SUM(B3:B14)" 
ws3['C15'] = "=SUM(C3:C14)"
ws3['D15'] = "=SUM(D3:D14)"
ws3['E15'] = "=SUM(E3:E14)"
ws3['F15'] = "=SUM(F3:F14)"
ws3['G15'] = "=SUM(G3:G14)"
ws3['H15'] = "=SUM(H3:H14)"
ws3['I15'] = "=SUM(I3:I14)"
ws3['J15'] = "=SUM(J3:J14)"
ws3['K15'] = "=SUM(K3:K14)"
ws3['L15'] = "=SUM(L3:L14)"
ws3['M15'] = "=SUM(M3:M14)"
ws3['N15'] = "=SUM(N3:N14)"
ws3['O15'] = "=SUM(O3:O14)"


# rounding the values in the summary sheet

for row in range(3,16):
    ws3[f'C{row}'].number_format = '0.00'
    ws3[f'E{row}'].number_format = '0.00'
    ws3[f'G{row}'].number_format = '0.00'
    ws3[f'I{row}'].number_format = '0.00'
    ws3[f'K{row}'].number_format = '0.00'
    ws3[f'M{row}'].number_format = '0.00'
    ws3[f'O{row}'].number_format = '0.00'
    

wb2.save(main_path) 
wb2.close()                       

print('ALL THE FORMATTING AND THE PUTTING VALUE IN THE SUMMARY SHEET HAS BEEN DONE')

print('NOW WORKING ON MAIL TO SEND')
sheet_name = 'SUMMARY'
summary_range = 'A1:O15'

subject_date = pd.Timestamp.now().strftime('%d%b%Y') 

to_mail = 'ravi.gupta@sbfc.com'
subject= f'HL-Summary as on {subject_date}'

excel = win32.Dispatch('Excel.Application')
excel.Visible=False
excel.DisplayAlerts = False

wb4 = excel.Workbooks.Open(main_path)
excel.CalculateFull()

time.sleep(1)

ws4 = wb4.Worksheets(sheet_name)
ws4.Range(summary_range).CopyPicture(Format = 2)

outlook = win32.Dispatch('Outlook.Application')
mail = outlook.CreateItem(0)

mail.to = to_mail
mail.Subject = subject

mail.HTMLBody = f"""
<html>
<body>
<p> Hello Team,</p>

<p> Please Find attached HL summary for Gujarat, branch wise - as on {subject_date}. </p>

<p> The Details include Logins | Approved | Subjective Approval | WIP | Disbursed | Declined.</p>
<br>
</body>
</html>
"""

mail.Display()

word_editor = mail.GetInspector.WordEditor
selection = word_editor.Application.Selection
selection.EndKey(6)

selection.Paste()

selection.TypeParagraph()
selection.TypeText('Regards,')
selection.TypeParagraph()
selection.TypeText('Ravi Gupta')

mail.Attachments.Add(main_path)

wb4.Close(SaveChanges= False)
excel.Quit()
mail.Send()

mail_send_time = pd.Timestamp.now().strftime('%H:%M:%S')
print(f"MAIL SEND SUCCESSFULLY AT - {mail_send_time}")

                                 
