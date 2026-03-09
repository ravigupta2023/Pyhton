import pandas as pd
import win32com.client as win
import pythoncom
import os
import shutil
import pandas as pd
import numpy as np
from datetime import timedelta
outlook = win.Dispatch('Outlook.Application').GetNameSpace('MAPI')
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items
messages.Sort('[ReceivedTime]',True)
date_for_outlook = pd.Timestamp.now().strftime('%d%b%Y')
# Delete_this = pd.Timestamp.now()
# yesterday_day_for_some = Delete_this-timedelta(days=1)
outlook_path = r"C:\Users\ET0001301\Pictures\Data\Sub Approved but not Approved\Outlook Data"
outlook_file = f'ME-MIS-{date_for_outlook}.xlsx'
outlook_file_save_path = os.path.join(outlook_path,outlook_file)
for mail in messages:
    sub = mail.Subject
    if 'ME-MIS-' in sub and date_for_outlook in sub:
        if mail.Attachments.count>=2:
            second_file = mail.Attachments.Item(2)
            second_file.SaveAsFile(outlook_file_save_path)
        mail_info = {
                    "sub":mail.subject,
                    'sender':mail.sender,
                    'time':mail.ReceivedTime
                    }    
        for k,v in mail_info.items():
            print(k,v)   
        break                 
shutil_path = r"C:\Users\ET0001301\Pictures\Data\Shutil Data"

os.makedirs(shutil_path,exist_ok=True)
# shutil_file = f'ME-MIS-{date_for_outlook}.xlsx'
shutil.copy(outlook_file_save_path,shutil_path)
main_path = os.path.join(shutil_path,outlook_file)
df = pd.read_excel(main_path,sheet_name='out_file')
df.insert(df.columns.get_loc('SCHEMETYPE')+1,'Scheme New',pd.NA)
df.insert(df.columns.get_loc('login_date')+1,'Dec_month',pd.NA)
schemes = ['HL','LAP','SBL/ML']
condtion1 = [df['SCHEMETYPE'].isin(['HL','HL BT','HL BT Top up','HL_ Topup -Parallel']),
             df['SCHEMETYPE'].isin(['BT','BT Topup','Normal','SALARIED LAP','Salaried LAP','SALARIED LAP BT TOPUP','SALARIED LAP BT','Top Up Gross','Top Up Parallel','Top-up-PreApproved']),
             df['SCHEMETYPE'].isin(['Micro loans','Micro loans BT','Micro_Topup-Parallel','SBL','SBL_BT','SBL_Top up-Parallel','SBL_Top up','Secured BL','Secured BL BT','Secured BL BT Top-up'])]
choice= ['HL','LAP','SBL/ML']
by = ''

df['Scheme New'] = np.select(condtion1,choice,by)
df['Scheme New']
df['login_date'] = pd.to_datetime(df['login_date'],errors = 'coerce')
current_month_name = pd.Timestamp.now().strftime('%b')
current_year = pd.Timestamp.now().year
current_month_number = pd.Timestamp.now().month

df['Dec_month'] = df['login_date'].apply(lambda x:current_month_name 
                                         if ((x.month == current_month_number) and (x.year == current_year))
                                         else "")
df = df[df['MIS_STATUS']=='Subjective Approval']
today_date_for_filter = pd.Timestamp.now()
last_2_days = today_date_for_filter-timedelta(days=2)
last_2_days
df = df[df['login_date']<=last_2_days]
df.shape
save_path = r"C:\Users\ET0001301\Pictures\Data\Sub Approved but not Approved"
os.makedirs(save_path,exist_ok=True)
file_name = 'test.xlsx'
save_path_final = os.path.join(save_path,file_name)
df.to_excel(save_path_final,index = False)
df1 = df.groupby(['ZONE','REGION','Scheme New'])['IN_Cr'].sum().unstack(fill_value = 0)
df1 = df1.reset_index()
df1.columns.name = None
df1
order = ['Bihar','Assam',
'Delhi',
'Haryana',
'Rajasthan',
'Uttarakhand',
'UP East',
'UP West',
'Varanasi',
'AP 1',
'AP 2',
'Bangalore',
'Mysore',
'Telangana 1',
'Telangana 2',
'TN',
'Mumbai',
'ROM',
'ROM 2',
'Gujarat',
'MP',
]
df1['REGION'] = pd.Categorical(df1['REGION'],categories=order,ordered=True)
df1 = df1.sort_values('REGION').reset_index(drop = True)
df1.columns
east_total = pd.DataFrame({'ZONE':['East'],'REGION':['East Total'],
                           'HL':[df1.loc[df1['REGION'].isin(['Assam',"Bihar"]),'HL'].sum()],
                           'LAP':[df1.loc[df1['REGION'].isin(['Assam','Bihar']),'LAP'].sum()],
                           'SBL/ML':[df1.loc[df1['REGION'].isin(['Assam','Bihar']),'SBL/ML'].sum()]})
pos = df1.loc[df1['ZONE']=='East'].index.max()+1
df1 = pd.concat([df1.iloc[:pos],east_total,df1.iloc[pos:]],ignore_index=True)
df1
Delhi_Total = pd.DataFrame({'ZONE':['North'],'REGION':['Delhi Total'],
                           'HL':[df1.loc[df1['REGION'].isin(['Delhi','Haryana','Uttarakhand','UTK','Uttrakhand','Rajasthan']),'HL'].sum()],
                           'LAP':[df1.loc[df1['REGION'].isin(['Delhi','Haryana','Uttarakhand','UTK','Uttrakhand','Rajasthan']),'LAP'].sum()],
                           'SBL/ML':[df1.loc[df1['REGION'].isin(['Delhi','Haryana','Uttarakhand','UTK','Uttrakhand','Rajasthan']),'SBL/ML'].sum()]})
utk_below = df1[df1['REGION']=="UTK"]
uttarakhand_below = df1[df1['REGION']=="Uttarakhand"]
delhi_below = df1[df1['REGION']=="Delhi"]
rj_below = df1[df1['REGION']=="Rajasthan"]
hy_below = df1[df1['REGION']=="Haryana"]

if not utk_below.empty:
    pos2 = utk_below.index.max()+1
elif not uttarakhand_below.empty:
    pos2 = uttarakhand_below.index.max()+1
elif not rj_below.empty:
    pos2 = rj_below.index.max()+1
elif not hy_below.empty:
    pos2 = hy_below.index.max()+1
elif not delhi_below.empty:
    pos2 = delhi_below.index.max()+1         

df1 = pd.concat([df1.iloc[:pos2],Delhi_Total,df1.iloc[pos2:]],ignore_index=True)
df1
up_total = pd.DataFrame({'ZONE':['North'],'REGION':['UP Total'],
                         'HL':[df1.loc[df1['REGION'].isin(['UP East','UP West','Varanasi']),'HL'].sum()],
                         'LAP':[df1.loc[df1['REGION'].isin(['UP East','UP West','Varanasi']),'LAP'].sum()],
                         'SBL/ML':[df1.loc[df1['REGION'].isin(['UP East','UP West','Varanasi']),'SBL/ML'].sum()],
                         })
vs_below = df1[df1['REGION']=="Varanasi"]
east_below  = df1[df1['REGION']=="UP East"]
west_below = df1[df1['REGION']=='UP West']
if not vs_below.empty:
    pos3 = vs_below.index.max()+1
elif not west_below.empty:
    pos3 = west_below.index.max()+1
elif not east_below.empty:
    pos3 = east_below.index.max()+1        
df1 = pd.concat([df1.iloc[:pos3],up_total,df1.iloc[pos3:]],ignore_index=True)
df1
North_total = pd.DataFrame({'ZONE':['North'],'REGION':['North Total'],
                            'HL':[df1.loc[df1['REGION'].isin(['Delhi Total','UP Total']),'HL'].sum()],
                            'LAP':[df1.loc[df1['REGION'].isin(['Delhi Total','UP Total']),'LAP'].sum()],
                            'SBL/ML':[df1.loc[df1['REGION'].isin(['Delhi Total','UP Total']),'SBL/ML'].sum()]})
pos4 = df1.loc[df1['ZONE']=='North'].index.max()+1
df1 = pd.concat([df1.iloc[:pos4],North_total,df1.iloc[pos4:]],ignore_index=True)
df1
ap_total = pd.DataFrame({'ZONE':['South'],
                         'REGION':['AP Total'],
                         'HL':[df1.loc[df1['REGION'].isin(['AP 1','AP 2']),'HL'].sum()],
                         'LAP':[df1.loc[df1['REGION'].isin(['AP 1','AP 2']),'LAP'].sum()],
                         'SBL/ML':[df1.loc[df1['REGION'].isin(['AP 1','AP 2']),'SBL/ML'].sum()]})
ap1_below = df1.loc[df1['REGION']=='AP 1']
ap2_below = df1.loc[df1['REGION']=='AP 2']
df1
if not ap2_below.empty:
    pos5 = ap2_below.index.max()+1
elif not ap1_below.empty:
    pos5 = ap1_below.index.max()+1    
try:
    if not ap1_below.empty or not ap2_below.empty:
        df1 = pd.concat([df1.iloc[:pos5],ap_total,df1.iloc[pos5:]],ignore_index = True)
except Exception as e:
    print('Nothing is present')
            
df1
karnataka_total = pd.DataFrame({'ZONE':['South'],
                                'REGION':['Karnataka Total'],
                                'HL':[df1.loc[df1['REGION'].isin(['Bangalore','Mysore']),'HL'].sum()],
                                'LAP':[df1.loc[df1['REGION'].isin(['Bangalore','Mysore']),'LAP'].sum()],
                                'SBL/ML':[df1.loc[df1['REGION'].isin(['Bangalore','Mysore']),'SBL/ML'].sum()]})
mysore_below = df1[df1['REGION']=='Mysore']
bang_below = df1[df1['REGION']=='Bangalore']
if not mysore_below.empty:
    pos6 = mysore_below.index.max()+1
elif not bang_below.empty:
    pos6 = bang_below.index.max()+1    

df1 = pd.concat([df1.iloc[:pos6],karnataka_total,df1.iloc[pos6:]],ignore_index=True)
df1
telangana_total = pd.DataFrame({'ZONE':['South'],
                                'REGION':['Telangana Total'],
                                'HL':[df1.loc[df1['REGION'].isin(['Telangana 1','Telangana 2']),'HL'].sum()],
                                'LAP':[df1.loc[df1['REGION'].isin(['Telangana 1','Telangana 2']),'LAP'].sum()],
                                'SBL/ML':[df1.loc[df1['REGION'].isin(['Telangana 1','Telangana 2']),'SBL/ML'].sum()]})
tn2_below = df1[df1['REGION']=='Telangana 2']
tn1_below = df1[df1['REGION']=='Telangana 1']
if not tn2_below.empty:
    pos7 = tn2_below.index.max()+1
elif not tn1_below.empty:
    pos7 = tn1_below.index.max()+1    
df1 = pd.concat([df1.iloc[:pos7],telangana_total,df1.iloc[pos7:]],ignore_index=True)
df1
South_total = pd.DataFrame({'ZONE':['South'],
                         'REGION':['South Total'],
                         'HL':[df1.loc[df1['REGION'].isin(['TN','Telangana Total','AP Total','Karnataka Total']),'HL'].sum()],
                         'LAP':[df1.loc[df1['REGION'].isin(['TN','Telangana Total','AP Total','Karnataka Total']),'LAP'].sum()],
                         'SBL/ML':[df1.loc[df1['REGION'].isin(['TN','Telangana Total','AP Total','Karnataka Total']),'SBL/ML'].sum()]})
pos8 = df1.loc[df1['ZONE']=='South'].index.max()+1
df1 = pd.concat([df1.iloc[:pos8],South_total,df1.iloc[pos8:]],ignore_index = True)
mh_total = pd.DataFrame({'ZONE':['West'],
                         'REGION':['Maharashtra Total'],
                         'HL':[df1.loc[df1['REGION'].isin(['ROM','ROM 2','Mumbai']),'HL'].sum()],
                         'LAP':[df1.loc[df1['REGION'].isin(['ROM','ROM 2','Mumbai']),'LAP'].sum()],
                         'SBL/ML':[df1.loc[df1['REGION'].isin(['ROM','ROM 2','Mumbai']),'SBL/ML'].sum()]})
rom2_below = df1[df1['REGION']=='ROM 2']
rom_below = df1[df1['REGION']=='ROM 1']
Mumbai_below = df1[df1['REGION']=='Mumbai']

if not rom2_below.empty:
    pos9 = rom2_below.index.max()+1
elif not rom_below.empty:
    pos9 = rom_below.index.max()+1
elif not Mumbai_below.empty:
    pos9 = Mumbai_below.index.max()+1        
df1 = pd.concat([df1.iloc[:pos9],mh_total,df1.iloc[pos9:]],ignore_index=True)
west_total = pd.DataFrame({'ZONE':['ZONE'],
                           'REGION':['West Total'],
                           'HL':[df1.loc[df1['REGION'].isin(['Gujarat','MP','Maharashtra Total']),'HL'].sum()],
                           'LAP':[df1.loc[df1['REGION'].isin(['Gujarat','MP','Maharashtra Total']),'LAP'].sum()],
                           'SBL/ML':[df1.loc[df1['REGION'].isin(['Gujarat','MP','Maharashtra Total']),'SBL/ML'].sum()]})
pos10 = df1[df1['ZONE']=='West'].index.max()+1
df1 = pd.concat([df1.iloc[:pos10],west_total,df1.iloc[pos10:]],ignore_index=True)
grand_total = pd.DataFrame({'ZONE':['Total'],
                            'REGION':['Grand Total'],
                            'HL':[df1.loc[df1['REGION'].isin(['West Total','North Total','South Total','East Total']),'HL'].sum()],
                            'LAP':[df1.loc[df1['REGION'].isin(['West Total','North Total','South Total','East Total']),'LAP'].sum()],
                            'SBL/ML':[df1.loc[df1['REGION'].isin(['West Total','North Total','South Total','East Total']),'SBL/ML'].sum()]})
df1 = df1.round(2)
date_for_file = pd.Timestamp.now().strftime('%d-%b-%Y')
df1['Grand Total'] = df1.sum(axis=1,numeric_only=True)
df1
row_grand_total = pd.DataFrame({'REGION':['Grand Total'],
                                'HL':[df1.loc[df1['REGION'].isin(['West Total','South Total','North Total','East Total']),'HL'].sum()],
                                'LAP':[df1.loc[df1['REGION'].isin(['West Total','South Total','North Total','East Total']),'LAP'].sum()],
                                'SBL/ML':[df1.loc[df1['REGION'].isin(['West Total','South Total','North Total','East Total']),'SBL/ML'].sum()],
                                'Grand Total':[df1.loc[df1['REGION'].isin(['West Total','South Total','North Total','East Total']),'Grand Total'].sum()]})
# df1.loc['Grand Total'] = df1[['HL','LAP','SBL/ML']].sum()
df1
try:
    if not df1.empty:
        df1 = pd.concat([df1,row_grand_total],ignore_index = True)
except Exception as e:
    print('THe data frame is not available')        
df1
df1 = df1.drop(columns = ['ZONE'])
save_path = r"C:\Users\ET0001301\Pictures\Data\Sub Approved but not Approved"
os.makedirs(save_path,exist_ok=True)
file_name = f'Sub approved but not approved {date_for_file}.xlsx'
final_path = os.path.join(save_path,file_name)
with pd.ExcelWriter(final_path,engine= 'openpyxl') as writer:
    df.to_excel(writer,sheet_name = 'Data',index =False)
    df1.to_excel(writer,sheet_name = 'Summary',index = False,startrow = 1)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font,Border,Alignment,Side
wb = load_workbook(final_path)
ws = wb['Summary']
ws['A1']='Sub Approved'
ws.merge_cells('A1:E1')
for cells in ws[1]:
    cells.alignment  = Alignment(horizontal='center',vertical='center')
    cells.font = Font(bold=True)
    cells.border = Border(left = Side(border_style = 'thin'),
                          right = Side(border_style ='thin'),
                          bottom=Side(border_style ='thin'),
                          top = Side(border_style='thin'))
zone_total = ['East Total','West Total','North Total','South Total','Grand Total']
region_total = ['Delhi Total','UP Total','AP Total','Karnataka Total','Telangana Total','Maharashtra Total']
for rows in ws.iter_rows(min_row= 1,max_row= ws.max_row, min_col= 1,max_col = ws.max_column):
    for cell in rows:
        if cell.value and (cell.value in zone_total or cell.value == 'REGION'):
            for cell in rows:
                cell.fill = PatternFill(start_color='A9CCE3',fill_type='solid')
                cell.font=Font(bold = True)
for row in ws.iter_rows(min_row = 2,max_row = ws.max_row,min_col = 1,max_col = ws.max_column):
    for cell in row:
        if cell.value and cell.value in region_total:
            for cell in row:
                cell.fill = PatternFill(start_color='FADBD8',fill_type='solid')
                cell.font = Font(bold=True)
for row in ws.iter_rows(min_row = 1,max_row = ws.max_row,min_col = 1,max_col = ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal ='center',vertical = 'center')
        cell.border = Border(top = Side(border_style='thin'),bottom = Side(border_style='thin'),
                                 left = Side(border_style='thin'),
                                 right = Side(border_style='thin'))
for rows in ws[1]:
    rows.fill = PatternFill(start_color='A9CCE3',fill_type='solid')
    rows.font = Font(bold = True)
    
for cols in ws.iter_cols(min_row = 2,max_row = ws.max_row,min_col = 1,max_col = ws.max_column):
        max_length = 0
        col_letter = cols[0].column_letter
        for cell in cols:
            if cell.value is not None:
                max_length = max(max_length,len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length+2                        
for cols in ws.iter_cols():
    for cell in cols:
        cell.alignment = Alignment(horizontal='center',vertical='center')
wb.save(final_path)
wb.close()
outlook1 = win.Dispatch('Outlook.Application')
mail = outlook1.CreateItem(0)
mail.to = 'ravi.gupta@sbfc.com'
mail.subject = 'Subj Approved but not Approved'
html_table = df1.to_html(index = False)
last_2_days1 = pd.Timestamp.now()
last_2_days2 = last_2_days1-timedelta(days=2)
last_2_days2 = last_2_days2.strftime('%d-%b')
mail.HTMLBODY = f''' 
<p>Hi All,<\p>

<p>Please find attached list of cases that were Subj Approved on or before {last_2_days2}  but not yet Approved.<\p>

<p>The Tranch cases is also included.<\p>

<p>Product wise Summary :<\p>

{html_table}

<p>Regards,<br>Ravi Gupta<\p>
'''
mail.Attachments.Add(final_path)
mail.Display()
mail.Send()
