import pandas as pd
import os
import numpy as np
from openpyxl.styles import Alignment,Border,Side,Font,PatternFill
from openpyxl import load_workbook

path = r"C:\Users\ET0001301\Desktop\Python\ME-MIS_SUMMARY.xlsx"
df = pd.read_excel(path,sheet_name = 'out_file')
df.columns = df.columns.str.strip()
df['Tranch'] = df['SCHEMETYPE'].apply(
                                        lambda x:'Yes' if x in
                                        {'BT Topup',
                                                'HL BT Top up',
                                                'Micro loans BT Top up',
                                                'SALARIED LAP BT TOPUP',
                                                'SBL_Top up',
                                                'Secured BL BT Top-up',
                                                'Top-up-PreApproved'
                                                } 
                                        else 'NO')

df['MIS_STATUS'] = df.apply(
                    lambda x:'WIP' 
                    if ((x['CURRENTSTATUS'] in {'Queue','Hold'}) and (x['MIS_STATUS']=='Approved'))
                    else x['MIS_STATUS'],
                    axis = 1
                    )
df['LAST_MASTER_REMARK'] = df['LAST_MASTER_REMARK'].fillna('No Remark')
df['login_date'] = pd.to_datetime(df['login_date'])
df['DOCUMENTRECEIVEDATCPADATE'] = pd.to_datetime(df['DOCUMENTRECEIVEDATCPADATE'])

df.insert(4,'Final Region',pd.NA)
df.insert(13,'Decision_month',pd.NA)
df.insert(15,'Cibil_bucketing',pd.NA)
df.insert(30,'Logins',pd.NA)

condition = [(df['APPLICANTCIBILSCORE']>=700),(df['APPLICANTCIBILSCORE']==-1)]
choices = ['>=700','-1']
df['Cibil_bucketing'] = np.select(condition,choices,default = '')

current_month_name = pd.Timestamp.now().month_name()
current_month = pd.Timestamp.now().month
current_year = pd.Timestamp.now().year
current_month_name2 = pd.Timestamp.now().strftime('%b')

df['Decision_month'] = df['login_date'].apply(
                            lambda x:current_month_name2 
                            if ((x.month_name()==current_month_name) and (x.year == current_year))
                            else ''
                            )
df['login_date'] = df['login_date'].dt.strftime('%d-%b-%Y')


df['Logins'] = df['DOCUMENTRECEIVEDATCPADATE'].apply(
                            lambda x:current_month_name2 
                            if ((x.month_name()==current_month_name) and (x.year == current_year))
                            else ''
                            )
df['DOCUMENTRECEIVEDATCPADATE'] = df['DOCUMENTRECEIVEDATCPADATE'].dt.strftime('%d-%b-%Y')

condition2 = [df['MIS_STATUS']=='APPROVED - ICICI',df['MIS_STATUS']=='DISBURSED - ICICI',df['MIS_STATUS']=='WIP - ICICI']
choices2 = ['Approved','Disbursed','WIP']

df['MIS_STATUS'] =np.select(condition2,choices2,default=df['MIS_STATUS']) 

# Code for ABM CBM
filter_for_logins = df[(df['Tranch']=='NO') &(df['Logins']!='')]
filter_for_Approval_Sub_Approval = df[(df['Tranch']=='NO')&
                                      (df['MIS_STATUS'].isin({'Approved','Subjective Approval'}) 
                                       &(df['Decision_month']!=''))
                                      ]

filter_for_Disb = df[(df['MIS_STATUS']=='Disbursed') 
                    &(df['Decision_month']!='')
                    ]
df['REGISTRATIONStage'] = df['REGISTRATIONStage'].str.replace('KO','K0')
pivot_for_logins = pd.pivot_table(filter_for_logins,
                                 index = 'REGISTRATIONStage',
                                 columns = 'MIS_STATUS',
                                 values = 'IN_Cr',
                                 aggfunc = 'count',
                                 margins = True,
                                 margins_name='Total'
                                 )
pivot_for_approval = pd.pivot_table(filter_for_Approval_Sub_Approval,
                                    index = 'REGISTRATIONStage',
                                    values = 'IN_Cr',
                                    aggfunc = ['count','sum'],
                                    margins = True,
                                    margins_name='Total'
                                    )
pivot_for_disb = pd.pivot_table(filter_for_Disb,
                                index = 'REGISTRATIONStage',
                                values = 'IN_Cr',
                                aggfunc = ['count','sum'],
                                margins = True,
                                margins_name='Total'
                                )

save_path = r'C:\Users\ET0001301\Desktop\Python'
file_name = 'data.xlsx'
final_path = os.path.join(save_path,file_name)
with pd.ExcelWriter(final_path,engine = 'openpyxl') as writer:
    df.to_excel(writer,sheet_name = 'Main Data',index = False)
    tables = [('Login_summary',pivot_for_logins),
              ('Approval Summary',pivot_for_approval),
              ('Disb Summary',pivot_for_disb)]
    row = 0
    col = 0
    for name,pivot in tables:
            ws = writer.book.create_sheet('ABM CBM') if 'ABM CBM' not in writer.book.sheetnames else writer.book['ABM CBM']
            ws.cell(row = row+1,column = col+1).value = name
            pivot.to_excel(writer,sheet_name = 'ABM CBM',startrow = row+2,startcol = col)
            col +=pivot.shape[1]+2

alignment = Alignment(horizontal='center',vertical='center')
header_fill = PatternFill(start_color='87CEEB',fill_type='solid')
font = Font(bold = True)
total_fill = PatternFill(start_color='87CEEB',fill_type='solid')

wb1 = load_workbook(final_path)
ws1 = wb1['ABM CBM']
ws2 = wb1['Main Data']
for col in ws1.iter_cols():
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
                if cell.value is not None:
                        max_length = max(max_length,len(str(cell.value)))
        ws1.column_dimensions[col_letter].width  = max_length+3
                                      
for cols in ws2.iter_cols():
        max_length = 0
        col = cols[0].column_letter
        for cell in cols:
                if cell.value is not None:
                        cell.alignment= alignment
                        max_length= max(max_length,len(str(cell.value)))
        ws2.column_dimensions[col].width = max_length+2                        
                                                 
wb1.save(final_path)
wb1.close()

# Decline data starting
decline_data = df[(df['Logins'] != '')&(df['MIS_STATUS']=='Declined') & (df['Cibil_bucketing'] != "")]
filter_for_east = decline_data[decline_data['ZONE'].str.lower()=='east']
filter_for_west = decline_data[decline_data['ZONE'].str.lower()=='west']
filter_for_north = decline_data[decline_data['ZONE'].str.lower()=='north']
filter_for_south = decline_data[decline_data['ZONE'].str.lower()=='south']

def declined_data(data):
        pivot = pd.pivot_table(data,
                                index = 'REGION',
                                columns = 'LAST_MASTER_REMARK',
                                values = 'REFERENCEID',
                                aggfunc= 'count',
                                margins = True,
                                margins_name='Total'
                                ).reset_index()
        pivot = pivot.fillna('-')
        return pivot 
east_pivot =declined_data(filter_for_east)           
west_pivot =declined_data(filter_for_west)           
north_pivot =declined_data(filter_for_north)           
south_pivot =declined_data(filter_for_south)

decline_data_path = r"C:\Users\ET0001301\Desktop\Python\Declined data"
os.makedirs(decline_data_path,exist_ok=True)
east_file = 'east_declined_data.xlsx'
west_file ='west_declined_data.xlsx'            
north_file = 'north_declined_data.xlsx'           
south_file = 'south_declined_data.xlsx'

def file_save_path(path,file_name):
        path = os.path.join(path,file_name)
        return path

east_file_save_path = file_save_path(decline_data_path,east_file)
west_file_save_path = file_save_path(decline_data_path,west_file)
north_file_save_path = file_save_path(decline_data_path,north_file)
south_file_save_path = file_save_path(decline_data_path,south_file)

filter_for_east.to_excel(east_file_save_path,sheet_name = 'Data',index = False)
filter_for_west.to_excel(west_file_save_path,sheet_name = 'Data',index = False)
filter_for_north.to_excel(north_file_save_path,sheet_name = 'Data',index = False)
filter_for_south.to_excel(south_file_save_path,sheet_name = 'Data',index = False)

with pd.ExcelWriter(east_file_save_path,engine='openpyxl',mode = 'a',if_sheet_exists='replace') as writer:
        east_pivot.to_excel(writer,sheet_name='Summary',index=False)
        
with pd.ExcelWriter(west_file_save_path,engine='openpyxl',mode = 'a',if_sheet_exists='replace') as writer:
        west_pivot.to_excel(writer,sheet_name='Summary',index=False)

with pd.ExcelWriter(north_file_save_path,engine='openpyxl',mode = 'a',if_sheet_exists='replace') as writer:
        north_pivot.to_excel(writer,sheet_name='Summary',index=False)

with pd.ExcelWriter(south_file_save_path,engine='openpyxl',mode = 'a',if_sheet_exists='replace') as writer:
        south_pivot.to_excel(writer,sheet_name='Summary',index=False)
         
# Decoline data formating kane ja rha ahu
side = Side(style = 'thin')
align = Alignment(horizontal='center',vertical='center')
boder = Border(left = side,
               right = side,
               top = side,
               bottom = side
               )
header = PatternFill(start_color='87CEEB',fill_type='solid')
total = PatternFill(start_color='87CEEB',fill_type='solid')
fon = Font(bold = True)
def format(path):
        db1 = load_workbook(path)
        ds1 = db1['Data']
        ds2 = db1['Summary']
        def data_column_width(ds1):
                for cols in ds1.iter_cols(min_row = 1,max_row = ds1.max_row,min_col = 1,max_col=ds1.max_column):
                        max_length = 0
                        col_letter = cols[0].column_letter
                        for cells in cols:
                                if cells.value is not None:
                                        max_length = max(max_length,len(str(cells.value)))
                        ds1.column_dimensions[col_letter].width = max_length+2                                               
        def summary_column_width(ds2):        
                for cols in ds2.iter_cols(min_row = 1,max_row = ds1.max_row,min_col = 1,max_col=ds1.max_column):
                        max_length = 0
                        col_letter = cols[0].column_letter
                        for cells in cols:
                                if cells.value is not None:
                                        max_length = max(max_length,len(str(cells.value)))
                        ds2.column_dimensions[col_letter].width = max_length+2                                               
        data_column_width(ds1)
        summary_column_width(ds2)
 # formatting eh sheet now
        def formatting(ds1):
                for cells in ds1[1]:
                        if cells.value is not None:
                                cell.fill = header
                                cell.border = boder
                                cell.alignment = align
                for cols in ds1.iter_rows():
                        for cells in cols:
                                cell.border = boder
                                cell.alignment  = align
                                
        formatting(ds1)
        def formatting_summary(ds2):
                for cell in ds2[1]:
                        if cell.value is not None:
                                cell.fill = header
                                cell.alignment = align
                                cell.border = boder
                                cell.font = fon
        formatting_summary(ds2)
        def formatting_an_sheet(ds2):
                for cols in ds2.iter_rows(min_row = 2,max_row =ds2.max_row,min_col = 1,max_col = ds2.max_column ):
                        for cell in cols:
                                if cell.value is not None:
                                        cell.border = boder
                                        cell.alignment= align
        formatting_an_sheet(ds2) 
        # Total Fomatting
        def total(ds2):
                for rows in ds2.iter_rows():
                        if rows[0].value == 'Total':
                                for cell in rows:
                                        cell.font = fon        
        total(ds2)      
        db1.save(path)
        db1.close()
                                                                  
east = format(east_file_save_path)
west = format(west_file_save_path)       
north = format(north_file_save_path)
south = format(south_file_save_path) 


# ab HL SUMMARY MAP KARNA HAi.

hl_path = r"C:\Users\ET0001301\Desktop\Python\ME-MIS_HL_SUMMARY.xlsx"
hl = pd.read_excel(hl_path,sheet_name = 'out_file')
gj = hl[hl['REGION'].str.lower()=='gujarat']


gj['Tranch'] = gj['SCHEMETYPE'].apply(lambda x:'YES' if x == 'HL BT Top up' else "NO")
gj['MIS_STATUS'] = gj.apply(lambda x:'WIP' if (
                                                (x['CURRENTSTATUS'] in {'Hold','Queue'}) and 
                                               (x['MIS_STATUS'] =='Approved')) else x['MIS_STATUS'],
                                                axis = 1)
gj.insert(12,'Decision_month',pd.NA)
gj.insert(28,'Logins',pd.NA)

gj['login_date'] = pd.to_datetime(gj['login_date'])
gj['DOCUMENTRECEIVEDATCPADATE'] = pd.to_datetime(gj['DOCUMENTRECEIVEDATCPADATE'])

short_month_name = pd.Timestamp.now().strftime('%b')
full_month = pd.Timestamp.now().month_name()
gj['Decision_month'] = gj['login_date'].apply(
                                                lambda x:short_month_name 
                                                if ((x.month_name()==full_month)) 
                                                else ''
                                                ) 

gj['Logins'] = gj['DOCUMENTRECEIVEDATCPADATE'].apply(
                                                lambda x:short_month_name 
                                                if ((x.month_name()==full_month)) 
                                                else ''
                                                ) 
with pd.ExcelWriter(final_path,engine = 'openpyxl',mode = 'a',if_sheet_exists='replace') as writer:
        gj.to_excel(writer,sheet_name = 'HL DATA',index = False)
# hl data bas save kiya hai


# Thoda sa chod kar ab Subjective Approval vala sreenivas sr k lia kr leta hu.

sa = df[(df['MIS_STATUS']=='Subjective Approval') & (df['Decision_month'] != '')]
with pd.ExcelWriter(final_path,engine= 'openpyxl',mode ='a',if_sheet_exists='replace') as writer:
        sa.to_excel(writer,sheet_name = 'Sub App Data',index = False)

sa_align = Alignment(horizontal='center',vertical='center')
# side = Side(style = 'thin')
# sa_border = Border(left = side,
#                    right = side,
#                    top = side,
#                    bottom = side)
sa1 = load_workbook(final_path)
sa2 = sa1['Sub App Data']

for cols in sa2.iter_cols():
        max_length = 0
        col_num = cols[0].column_letter
        for cell in cols:
                if cell.value is not None:
                        cell.alignment= sa_align
                        max_length = max(max_length,len(str(cell.value)))
        sa2.column_dimensions[col_num].width = max_length+2                        
# for cols in sa2.iter_cols(min_row = 1,max_row= sa2.max_row,min_col = 1,max_col = sa2.max_column):
#         for cell in cols:
#                 cell.border = sa_border
sa1.save(final_path)
sa1.close()
# Subjective Approval has been finished

# Sub App Less than 3 days cases.
from datetime import timedelta
today = pd.Timestamp.now().normalize()
excelude_days = 2

df['login_date'] = pd.to_datetime(df['login_date'])
three_days_ago = today-timedelta(excelude_days)
app_sub_3 = df[(df['MIS_STATUS'].isin(['Approved','Subjective Approval'])) & 
               (df['login_date'].dt.normalize() < three_days_ago) &
               (df['Tranch']=='NO')]


manish_path = r"C:\Users\ET0001301\Desktop\Python\MANISH SAPKAL.xlsx"               #manish sapkal file
manish_file = pd.read_excel(manish_path,sheet_name='DATA')

app_sub_3 = app_sub_3[~app_sub_3['REFERENCEID'].isin(manish_file['LeviosaID#'])]

three_days_pivot = pd.pivot_table(app_sub_3,
                                  values = 'IN_Cr',
                                  index = 'REGION',
                                  columns = 'MIS_STATUS',
                                  aggfunc =['count','sum'],
                                  margins = True,
                                  margins_name = "Total"
                                  ).fillna(0)
with pd.ExcelWriter(final_path,engine = 'openpyxl',mode = 'a',if_sheet_exists='replace') as writer:
        app_sub_3.to_excel(writer,sheet_name = 'App Sub less than 3 days',index = False)
        three_days_pivot.to_excel(writer,sheet_name='<3 Pivot')
        
man = load_workbook(final_path)
mans = man['App Sub less than 3 days']
mandsp = man['<3 Pivot']

malign = Alignment(horizontal='center',vertical='center')

for cols in mans.iter_cols(min_row = 1,max_row = mans.max_row,min_col = 1,max_col = mans.max_column):
        for cell in cols:
                cell.alignment = malign

for cols in mandsp.iter_cols(min_row = 1,max_row = mans.max_row,min_col = 1,max_col = mans.max_column):
        for cell in cols:
                cell.alignment = malign
                
for cols in mans.iter_cols(min_row = 1,max_row = mans.max_row,min_col = 1,max_col = mans.max_column):
        max_length = 0
        col_num = cols[0].column_letter
        for cell in cols:
                if cell.value is not None:
                        max_length=max(max_length,len(str(cell.value)))
        mans.column_dimensions[col_num].width = max_length + 2                        

from openpyxl.utils import get_column_letter        
for cols in mandsp.iter_cols():
        max_length = 0
        col_index = cols[0].column
        col_letter = get_column_letter(col_index)
        for cell in cols:
                if cell.value is not None:
                        max_length=max(max_length,len(str(cell.value)))
        mandsp.column_dimensions[col_letter].width = max_length + 2                        
                                           
man.save(final_path)
man.close()                                                                     

# Sumidha mam ka kam karne ja rha hu

sumida = r"C:\Users\ET0001301\Desktop\Python\DailyClosurePivot.xlsm"
closure = pd.read_excel(sumida,sheet_name = 'Sheet1')

me_closure = closure[(closure['ProductCheck']=='ME') & 
                     (closure['REASON CODE'].isin(['NORMAL FORECLOSURE','CANCELLATION']))
                     ]
clsoure_pivot = pd.pivot_table(me_closure,
                               values = 'IN_CR',
                               index = 'Region2',
                               columns = 'REASON CODE',
                               aggfunc='sum',
                               margins=True,
                               margins_name='Total'
                               ).fillna(0)
with pd.ExcelWriter(final_path,engine='openpyxl',mode = 'a',if_sheet_exists='replace') as writer:
        me_closure.to_excel(writer,sheet_name='Closure Data',index = False)
        clsoure_pivot.to_excel(writer,sheet_name='Closure Pivot')


                                                                                
                                                                                        


