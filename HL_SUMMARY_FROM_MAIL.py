import win32com.client as win32
import os 
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side,Alignment,PatternFill
from openpyxl.utils import get_column_letter
import time

save_path = r"C:\Users\et0001301\Documents"
file_name = 'HL-Summary.xlsx'

final_path = os.path.join(save_path,file_name)
# fil_name = "ME_HL DATA CHECK.xlsx"

# final_path = os.path.join(save_path,fil_name)

today_str = datetime.now().strftime('%d%b%Y')


outlook = win32.Dispatch('Outlook.Application').GetNamespace('MAPI')

inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items

messages.Sort("[ReceivedTime]", True)
file_found = False

for mail in messages:
    subject = mail.Subject
    
    if subject.startswith('ME-MIS-') and today_str in subject:
        if mail.Attachments.count>=2:
            if os.path.exists(final_path):
                os.remove(final_path)
            second_attachment = mail.Attachments.Item(2)
            second_attachment.SaveAsFile(final_path)
            print("2nd attachment save success")

            mail_info = {
                "Subject": mail.Subject,
                "Sender": mail.SenderName,
                "ReceivedTime": mail.ReceivedTime,
                # "EntryID": mail.EntryID
            }

            print("Mail used to save the file:")
            for k, v in mail_info.items():
                print(f"{k}: {v}")
            break   

time.sleep(3)


df = pd.read_excel(final_path,sheet_name= 'out_file')

df = df[df['REGION']=='Gujarat']

df.loc[df['SCHEMETYPE']=='HL BT Top up','Tranch']='yes'

df['login_date']=pd.to_datetime(df['login_date'])

current_date = pd.Timestamp.now()

df['Decision_month']=df['login_date'].apply(lambda x:pd.Timestamp.now().strftime('%b') if (x.month_name() == current_date.month_name() and x.year == current_date.year) else "") 

df['DOCUMENTRECEIVEDATCPADATE']=pd.to_datetime(df['DOCUMENTRECEIVEDATCPADATE'])

df['logins']=df['DOCUMENTRECEIVEDATCPADATE'].apply(lambda x:pd.Timestamp.now().strftime('%b') if (x.month_name() == current_date.month_name() and x.year == current_date.year) else "")

current_month_name = pd.Timestamp.now()
short_month_name = current_month_name.strftime('%b')

filter_for_logins = df[(df['logins']== short_month_name) & (df['Tranch']=='NO')]

filter_for_approval = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='Approved') & (df['Tranch']=='NO')]

filter_for_sub_approval = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='Subjective Approval') & (df['Tranch']=='NO')]

filter_for_wip = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='WIP') & (df['Tranch']=='NO')]

filter_for_disbursed = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='Disbursed') & (df['Tranch'].isin({'yes','NO'}))]

filter_for_declined = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='Declined') & (df['Tranch'].isin({'yes','NO'}))]

def create_pivot(filtered_data):
    pivot = pd.pivot_table(filtered_data,
                           values = 'IN_Cr',
                           index = 'BRANCH',
                           aggfunc = ['count','sum'],
                           fill_value = 0,
                           margins = True,
                           margins_name = 'Total'
                           )
    return pivot
    
login_pivot = create_pivot(filter_for_logins)
approval_pivot = create_pivot(filter_for_approval)
sub_approval_pivot = create_pivot(filter_for_sub_approval)    
wip_pivot = create_pivot(filter_for_wip)
disbursed_pivot = create_pivot(filter_for_disbursed)
declined_pivot = create_pivot(filter_for_declined)

save_path = r"C:\Users\et0001301\Documents\Automatting"
os.makedirs(save_path,exist_ok = True )
file_name = 'HL_Summary.xlsx'

final_path = os.path.join(save_path,file_name) 

with pd.ExcelWriter(final_path,engine = 'openpyxl',mode = 'w') as writer:
    df.to_excel(writer,sheet_name = 'cleaned_data',index = False)
    
    tables = [('Logins',login_pivot),
              ('Approved_summary',approval_pivot),
              ('Sub_app_summary',sub_approval_pivot),
              ('WIP_Summary',wip_pivot),
              ('Disbursed',disbursed_pivot),
              ('Declined',declined_pivot)
              ]
            
    row = 0
    col = 0 
    for title,pivots in tables:
        ws = writer.book.create_sheet('pivots') if 'pivots' not in writer.book.sheetnames else writer.book['pivots']
        ws.cell(row = row+1,column = col+1).value = title
        ws.cell(row = row+1,column = col+1).font = Font(bold = True) 
        pivots.to_excel(writer,sheet_name = 'pivots',startrow = row + 2,startcol = col )
        col += pivots.shape[1]+2
                    
wb = load_workbook(final_path)
ws = wb['pivots']

font = Font(bold = True)
Alignments1 = Alignment(horizontal = 'center',vertical = 'center')
Header_Fill = PatternFill(start_color = 'ADD8E6',fill_type = 'solid')
total_fill = PatternFill(start_color = 'FFFF00',fill_type = 'solid')
thin_border = Border(left = Side(style = 'thin'),
                     right = Side(style = 'thin'),
                     top = Side(style = 'thin'),
                     bottom = Side(style = 'thin'))
# total_fill = PatternFill(start_color = '')

for cell in ws[1]:
    if cell.value is not None:
        cell.font = font
        cell.fill = Header_Fill
 
 
for row in ws.iter_rows():
    for cell in row:
        if cell.value == 'Total':
            cell.fill = total_fill
    
    
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            cell.border = thin_border         
            cell.alignment = Alignments1

for col in ws.iter_cols():
    max_length = 0
    column_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length,len(str(cell.value)))
    ws.column_dimensions[column_letter].width= max_length + 3
wb.save(final_path)   

wb1 = load_workbook(final_path) 
ws1 = wb1['cleaned_data']

for col in ws1.iter_cols():
    max_length = 0
    column_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length,len(str(cell.value)))
    ws1.column_dimensions[column_letter].width = max_length + 4 
    
for rows in ws1.iter_rows():
    for cell in rows:
        cell.border = thin_border
        cell.alignment = Alignments1
 
for rows in ws1[1]:
    rows.font = font

wb1.save(final_path)                
wb1.close()

from openpyxl.utils import get_column_letter

wb2 = load_workbook(final_path)
ws2 = wb2['pivots']

max_col = ws2.max_column

for col in range(2,max_col + 1):
    col_letter = get_column_letter(col)
    
    top = ws2[f"{col_letter}3"].value
    bottom = ws2[f"{col_letter}4"].value
    
    if top and bottom:
        ws2[f"{col_letter}3"].value = f"{top} {bottom}"
        ws2.merge_cells(f"{col_letter}3:{col_letter}4")

for col in ws2.iter_cols():
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length,len(str(cell.value)))
    ws2.column_dimensions[col_letter].width = max_length+3    

value_columns = 2
    
for rows in ws2.iter_rows():
    for idx,cell in enumerate(rows):
        if cell.value is not None and str(cell.value).strip().lower() == 'total':
            cell.fill = total_fill
            
            for rcell in rows[idx+1:idx+1+value_columns]:
                if rcell.value is not None:
                    rcell.fill = total_fill     
wb2.save(final_path)
wb2.close()

wb3 = load_workbook(final_path)
ws3 = wb3.create_sheet('Summary')

main_headers = ['Branch','Logins','Approved','Sub Approval','WIP','Disbursed','Declined','Total']
sub_header = ['count','sum']

column_width = [15,8,8,8,8,8,8,8,8,8,8,8,8,8,8]

for i,width in enumerate(column_width,1):
    col_letter = get_column_letter(i)
    ws3.column_dimensions[col_letter].width = width

col = 1
for header in main_headers:
    if header == 'Branch':
            ws3.cell(row = 1,column = 1,value = header)
            ws3.merge_cells(start_row = 1,end_row=2,start_column = col,end_column = col)
            col += 1
    else:
        ws3.cell(row = 1,column = col,value = header)
        ws3.merge_cells(start_row = 1,start_column = col,end_row = 1,end_column = col+1)
        col += 2

col = 2        
for _ in main_headers[1:]:
    ws3.cell(row = 2,column = col,value = 'count')
    ws3.cell(row = 2,column = col+1,value = 'sum')
    col += 2

branches_consolidated = ['Anand','Gurukul','Maninagar','Mavdi','Mehsana','Morbi','Palanpur','Rajkot','Silvassa','Surat','Vadodara','Vapi','Total']

row = 3
col = 1
for branch in branches_consolidated:
    ws3.cell(row = row,column =col,value = branch )
    row +=1    

alignment = Alignment(horizontal = 'center',vertical = 'center')
border = Border(left = Side(style = 'thin'),
                right = Side(style = 'thin'),
                top = Side(style = 'thin'),
                bottom = Side(style = 'thin')
                )
branch_colour = PatternFill(start_color = 'ADD8E6',fill_type = 'solid')                 
other_header_colour = PatternFill(start_color = 'FFB6C1',fill_type = 'solid')
font = Font(bold = True)

for row in ws3.iter_rows(min_row = 1,max_row = 2,min_col = 1,max_col=1):
    for cell in row:
        cell.fill = branch_colour
        cell.font = font

for row in ws3.iter_rows(min_row = 1,max_row = 2,min_col = 2,max_col = 15):
    for cell in row:
        cell.fill = other_header_colour
        cell.alignment = alignment
        cell.font = font

for row in ws3.iter_rows(min_row = 1,max_row = 15,min_col = 1,max_col = 1):
    for cell in row:
        if cell.value == 'Total':
            cell.fill = branch_colour
            cell.font = font

for row in ws3.iter_rows(min_row = 1,max_row = 15,min_col = 1,max_col = 15):
    for cell in row:
        cell.border = border
        cell.alignment = alignment                            

wb3.save(final_path)
wb3.close()

wb4 = load_workbook(final_path)
ws_pivot = wb4['pivots']
ws_summary = wb4['Summary']

ws_summary['N3']="=sum(D3,F3,J3)"
for row in range(3,15):
    ws_summary[f'N{row}']=f"=sum(D{row},F{row},J{row})"
        
ws_summary['O3']="round(=sum(E3,G3,K3),2)"
for row in range(3,15):
    ws_summary[f'O{row}']=f"=sum(E{row},G{row},K{row})"    
    # ws_summary[f"O{row}"]=round(ws_summary[f"O{row}"],2)
    
ws_summary['B3']="=SUMIF(pivots!A:A,Summary!A3,pivots!B:B)"
ws_summary['C3']="=ROUND(SUMIF(pivots!A:A,Summary!A3,pivots!C:C),2)"

ws_summary['D3']='=SUMIF(pivots!E:E,Summary!A3,pivots!F:F)'
ws_summary['E3']="=round(SUMIF(pivots!E:E,Summary!A3,pivots!G:G),2)"

ws_summary['F3']="=SUMIF(pivots!I:I,Summary!A3,pivots!J:J)"     
ws_summary['G3']="=round(SUMIF(pivots!I:I,Summary!A3,pivots!K:K),2)"

ws_summary['H3']="=SUMIF(pivots!M:M,Summary!A3,pivots!N:N)"
ws_summary['I3']="=round(SUMIF(pivots!M:M,Summary!A3,pivots!O:O),2)"

ws_summary['J3']="=SUMIF(pivots!Q:Q,Summary!A3,pivots!R:R)"
ws_summary['K3']="=round(SUMIF(pivots!Q:Q,Summary!A3,pivots!S:S),2)"

ws_summary['L3']="=SUMIF(pivots!U:U,Summary!A3,pivots!V:V)"
ws_summary['M3']="=round(SUMIF(pivots!U:U,Summary!A3,pivots!W:W),2)"


for row in range(3,15):
    ws_summary[f"B{row}"]=ws_summary['B3'].value.replace('A3',f'A{row}')
    ws_summary[f'C{row}']=ws_summary['C3'].value.replace('A3',f'A{row}')
    ws_summary[f'D{row}']=ws_summary['D3'].value.replace('A3',f'A{row}')
    ws_summary[f'E{row}']=ws_summary['E3'].value.replace('A3',f'A{row}')
    ws_summary[f'F{row}']=ws_summary['F3'].value.replace('A3',f'A{row}')
    ws_summary[f'G{row}']=ws_summary['G3'].value.replace('A3',f'A{row}')
    ws_summary[f'H{row}']=ws_summary['H3'].value.replace('A3',f'A{row}')
    ws_summary[f'I{row}']=ws_summary['I3'].value.replace('A3',f'A{row}')
    ws_summary[f'J{row}']=ws_summary['J3'].value.replace('A3',f'A{row}')
    ws_summary[f'K{row}']=ws_summary['K3'].value.replace('A3',f'A{row}')
    ws_summary[f'L{row}']=ws_summary['L3'].value.replace('A3',f'A{row}')
    ws_summary[f'M{row}']=ws_summary['M3'].value.replace('A3',f'A{row}')

# column total Sum "Total"
ws_summary['B15']="=sum(B3:B14)"
ws_summary['C15']="=round(sum(C3:C14),2)"
ws_summary['D15']="=sum(D3:D14)"
ws_summary['E15']="=round(sum(E3:E14),2)"
ws_summary['F15']="=sum(F3:F14)"
ws_summary['G15']="=round(sum(G3:G14),2)"
ws_summary['H15']="=sum(H3:H14)"
ws_summary['I15']="=round(sum(I3:I14),2)"
ws_summary['J15']="=sum(J3:J14)"
ws_summary['K15']="=round(sum(K3:K14),2)"
ws_summary['L15']="=sum(L3:L14)"
ws_summary['M15']="=round(sum(M3:M14),2)"
ws_summary['N15']="=sum(N3:N14)"
ws_summary['O15']="=round(sum(O3:O14),2)"

for row in range(1,16):
    ws_summary[f'C{row}'].number_format ="0.00"
    
for row in range(1,16):
    ws_summary[f'E{row}'].number_format = "0.00"    
    
for row in range(1,16):
    ws_summary[f'G{row}'].number_format = "0.00"    

for row in range(1,16):
    ws_summary[f'I{row}'].number_format = "0.00"    

for row in range(1,16):
    ws_summary[f'K{row}'].number_format = "0.00"    

for row in range(1,16):
    ws_summary[f'M{row}'].number_format = "0.00"    

for row in range(1,16):
    ws_summary[f'O{row}'].number_format = "0.00"    

font = Font(bold = True)

for row in ws_summary.iter_rows():
    for idx,cell in enumerate(row,start=1 ):
        if str(cell.value).lower() == 'total':
            cell.font = font
            for rcell in row[idx:idx+14]:
                rcell.font = font
            
wb4.save(final_path)
wb4.close()

wb5 = load_workbook(final_path)
ws5 = wb5['Summary']

total_value_beside_color = PatternFill(start_color = 'FFB6C1',fill_type = 'solid')

for col in ws5.iter_cols(min_col = 1,max_col = 15,min_row = 15,max_row = 15):
    for idx,cell in enumerate(col,start=1):
        if cell.value == 'Total':
            start_col = cell.column
            for c in range (start_col+1,start_col+15):
                cells = ws5.cell(row = 15,column =c )
                cells.fill = total_value_beside_color

wb5.save(final_path)
wb5.close()

import time
time.sleep(5)
        
print("save successfull")

import win32com.client as win32

file_path = final_path

sheet_name = "Summary"
summary_range = 'A1:O15'

to_mail = 'ravi.gupta@sbfc.com'
subject = "HL-Summary as on 17th Dec'25"

excel = win32.Dispatch('Excel.Application')
excel.Visible = False
excel.DisplayAlerts = False

wb6 = excel.Workbooks.Open(file_path)
excel.CalculateFull()
time.sleep(2)

ws6 = wb6.Worksheets(sheet_name)

ws6.Range(summary_range).CopyPicture(Format = 2)

outlook = win32.Dispatch('Outlook.Application')
mail = outlook.CreateItem(0)

mail.to = to_mail
mail.Subject = subject

mail.HTMLBody = """
<html>
<body>
<p> Hello Team,</p>

<p> Please Find attached HL summary for Gujarat, branch wise - as on 17th Dec'25. </p>

<p> The Details include Logins | Approved | Subjective Approval | WIP | Disbursed | Declined.</p>
<br>
</body>
</html>
"""

mail.Display()

word_editor = mail.GetInspector.WordEditor
selection = word_editor.Application.Selection
selection.EndKey(6)

selection.Paste()

selection.TypeParagraph()
selection.TypeText('Regards,')
selection.TypeParagraph()
selection.TypeText('Ravi Gupta')

mail.Attachments.Add(file_path)

wb6.Close(SaveChanges= False)
excel.Quit()
mail.Send()


print("successfully send")

