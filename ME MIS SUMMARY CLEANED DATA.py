import pandas as pd
import os
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side,PatternFill,Alignment
import win32com.client as win32
import shutil

outlook = win32.Dispatch('Outlook.Application').GetNamespace('MAPI')
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items
messages.sort("[ReceivedTime]",True)

sub_current_date = pd.Timestamp.now().strftime('%d%b%Y')

outlook_path = r"C:\Users\et0001301\Desktop\Python\OUTLOOK_PATH"
outlook_data_file_name = 'outlook_me_mis.xlsx'

os.makedirs(outlook_path,exist_ok = True)
outlook_final_path = os.path.join(outlook_path,outlook_data_file_name)

for mail in messages:
    subject = mail.subject
    if 'ME-MIS-' in subject and sub_current_date in subject:
        if mail.Attachments.count>=2:
            second_attachment = mail.Attachments.Item(2)
            second_attachment.SaveAsFile(outlook_final_path)
        mail_info = {
                    "subject":mail.subject,
                    "sender":mail.sender,
                    'Received_time':mail.ReceivedTime
                    }       
        for k,v in mail_info.items():
            print(f'{k}:{v}')
        break         

path = r"C:\Users\et0001301\Desktop\Python\MAIN FOLDER"

os.makedirs(path,exist_ok = True)
shutil.copy(outlook_final_path,path)

df = pd.read_excel(outlook_final_path,sheet_name = 'out_file')

df.columns = df.columns.str.strip()

df.insert(12,'Decision_month',pd.NA)
df.insert(28,'Logins',pd.NA)

df['Tranch'] = df.apply(lambda x:'YES' if (
                                            x['SCHEMETYPE'] in {'BT Topup',
                                                'HL BT Top up',
                                                'Micro loans BT Top up',
                                                'SALARIED LAP BT TOPUP',
                                                'SBL_Top up',
                                                'Secured BL BT Top-up',
                                                'Top-up-PreApproved'
                                                })
                                        else "NO",
                                        axis = 1
                                        ) 

df['MIS_STATUS'] = df['MIS_STATUS'].apply(lambda x:'Approved' if x == 'APPROVED - ICICI' else ('Disbursed' if x=='DISBURSED - ICICI' else('WIP' if x == 'WIP - ICICI' else x)))

df['MIS_STATUS'] = df.apply(lambda x:'WIP' if ((x['CURRENTSTATUS'] in {'Hold','Queue'}) & (x['MIS_STATUS']=='Approved')) else x['MIS_STATUS'], axis = 1)

df['DOCUMENTRECEIVEDATCPADATE']= pd.to_datetime(df['DOCUMENTRECEIVEDATCPADATE'])
df['login_date'] = pd.to_datetime(df['login_date'])


current_date = pd.Timestamp.now()
current_year = current_date.year
current_month = current_date.month
current_month_name = current_date.strftime('%b')

df['Logins']=df['DOCUMENTRECEIVEDATCPADATE'].apply(lambda x:current_month_name if ((x.month == current_month) and (x.year == current_year)) else "")
df['Decision_month'] = df['login_date'].apply(lambda x:current_month_name if ((x.month == current_month) and (x.year == current_year)) else "")

df['login_date'] = df['login_date'].dt.strftime('%d-%b-%Y')
df['DOCUMENTRECEIVEDATCPADATE'] = df['DOCUMENTRECEIVEDATCPADATE'].dt.strftime('%d-%b-%Y')

df['REGISTRATIONStage'] = df['REGISTRATIONStage'].str.replace('KO','K0',case = False)

# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot
# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot
# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot
# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot

filter_for_logins = df[(df['Tranch'].str.lower()=="no") & (df['Logins'].str.lower()==current_month_name.lower())]
filter_for_approval = df[(df['Tranch'].str.lower()=='no') & (df['Decision_month'].str.lower()==current_month_name.lower()) & (df['MIS_STATUS'].isin(['Approved','Subjective Approval']))]
filter_for_disbursed = df[(df['Decision_month'].str.lower()==current_month_name.lower()) & (df['MIS_STATUS'].str.lower()=='Disbursed'.lower())]


login_pivot = pd.pivot_table(
                        filter_for_logins,
                        index = 'REGISTRATIONStage',
                        columns = 'MIS_STATUS',
                        values = 'IN_Cr',
                        aggfunc = 'count',
                        fill_value = 0,
                        margins = True,
                        margins_name = 'Grand Total'
                        )

approval_pivot = pd.pivot_table(
                                filter_for_approval,
                                index = 'REGISTRATIONStage',
                                values = 'IN_Cr',
                                aggfunc = ['count','sum'],
                                fill_value = 0,
                                margins = True,
                                margins_name = 'Grand Total'
                                )
disbursal_pivot = pd.pivot_table(
                                filter_for_disbursed,
                                index = 'REGISTRATIONStage',
                                values = 'IN_Cr',
                                aggfunc = ['count','sum'],
                                fill_value = 0,
                                margins = True,
                                margins_name = 'Grand Total'
                                )


save_path = r"C:\Users\et0001301\Desktop\Python\BACKUP FOLDER"
file_name = 'ME-MIS_Cleaned_data.xlsx'
os.makedirs(save_path,exist_ok = True)
final_path = os.path.join(save_path,file_name)

with pd.ExcelWriter(final_path,engine = 'openpyxl') as writer:
    df.to_excel(writer,sheet_name = 'Data',index = False)
    
    tables = [
            ('Login_summary',login_pivot),
            ('App_summary',approval_pivot),
            ('Disbursal_summary',disbursal_pivot)
    ]
    
    # if 'pivots' not in writer.book.sheetnames:
    #     ws1 = writer.book.create_sheet('pivots')
    # else:
    #     ws1 = writer['pivots']    
    
    col = 0
    row = 0
    for titles,pivots in tables:
        ws1 = writer.book.create_sheet('pivots') if 'pivots' not in writer.book.sheetnames else writer.book['pivots']
        ws1.cell(row =row+1,column= col+1).value = titles
        pivots.to_excel(writer,sheet_name = 'pivots',startrow = row+1,startcol =col )
        col +=pivots.shape[1]+2

time.sleep(3)        

wb = load_workbook(final_path)
ws1 = wb['Data']
ws2 = wb['pivots']

font_header = Font(bold = True,name = 'calibri')
font_value = Font(name = 'calibri')
header_color = PatternFill(start_color = '87CEEB',fill_type = 'solid')
pivot_header = PatternFill(start_color = '87CEEB',fill_type = 'solid')
border = Border(
                left = Side(style = 'thin'),
                right = Side(style = 'thin'),
                top = Side(style = 'thin'),
                bottom = Side(style = 'thin')
                )
alignment = Alignment(horizontal = 'center',vertical ='center')

for row in ws1.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = alignment
        cell.font = font_value
for cell in ws1[1]:
    cell.fill = header_color
    cell.font = font_header
        
for col in ws1.iter_cols(min_row = 1,max_row = ws1.max_row,min_col=1,max_col = ws1.max_column):
    for cell in col:
        max_length = 0
        col_letter = col[0].column_letter
        if cell.value is not None:
            max_length = max(max_length,len(str(cell.value)))
    ws1.column_dimensions[col_letter].width = max_length + 4                         
wb.save(final_path)
wb.close()

current_date_for_mail = pd.Timestamp.now().strftime('%d%b%Y')

send_outlook = win32.Dispatch('Outlook.Application')
mail = send_outlook.CreateItem(0)
mail.to = 'ravi.gupta@sbfc.com'
mail.Subject = f'ME-MIS-Summary-{current_date_for_mail}'
mail.HTMLBody = f"""
<html>
<body>
<p> 'Hello team,</p>

<p> Please find  attached ME-MIS-Summary as on {current_date_for_mail}.</p>

<p> Regards,</p>

<p> Ravi Gupta</p>

</body>
</html>
"""

mail.Attachments.Add(final_path)
mail.Send() 
        
        























