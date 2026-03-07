import win32com.client as win32
import pythoncom
import pandas as pd
import os
import shutil
from datetime import timedelta
pythoncom.CoInitialize()
outlook = win32.Dispatch('Outlook.Application').GetNameSpace('MAPI')
inbox = outlook.GetDefaultFolder(6)
messages = inbox.items
messages.Sort("[ReceivedTime]",True)
current_date = pd.Timestamp.now().strftime('%d%b%Y')
outlook = r"C:\Users\ET0001301\Pictures\Data\Approved not Disb\Outlook Data"
os.makedirs(outlook,exist_ok=True)
outlook_file_name = f'ME-MIS-{current_date}.xlsx'
outlook_path = os.path.join(outlook,outlook_file_name)
for mail in messages:
    sub = mail.Subject
    if 'ME-MIS-' in sub and current_date in sub:
        if mail.Attachments.count>=2:
            second_attachment = mail.Attachments.Item(2)
            second_attachment.SaveAsFile(outlook_path)
        mail_info = {
                    "sub":mail.subject,
                    'sender':mail.sender,
                    'Received Time':mail.ReceivedTime
                    }
        for k,v in mail_info.items():
            print(f'{k}:{v}')
        break    
shutil_path = r"C:\Users\ET0001301\Pictures\Data\Shutil Data"
os.makedirs(shutil_path,exist_ok=True)
file_name = outlook_file_name
shutil.copy(outlook_path,shutil_path)
main_path = os.path.join(shutil_path,outlook_file_name)
df = pd.read_excel(main_path,sheet_name='out_file')
df['MIS_STATUS']= df.apply(lambda x:'WIP' if ((x['MIS_STATUS']=='Approved') and 
                           (x['CURRENTSTATUS'] in ['Hold','Queue'])) else x['MIS_STATUS'],axis = 1)
df['login_date'] = pd.to_datetime(df['login_date'],errors = 'coerce')
df.insert(df.columns.get_loc('login_date')+1,'Dec_month',pd.NA)
current_month_name = pd.Timestamp.now().strftime('%b')
current_year = pd.Timestamp.now().year
current_month = pd.Timestamp.now().month
df['Dec_month'] = df['login_date'].apply(lambda x:current_month_name if ((x.month ==current_month) and (x.year == current_year))
                                         else "" )
df['MIS_STATUS'] = df['MIS_STATUS'].apply(lambda x:'Approved' if x=='APPROVED - ICICI' else x )
today_date_for = pd.Timestamp.now()
last_3_days = today_date_for-timedelta(days=3)
df = df[df['login_date']<=last_3_days]
df['login_date'] = df['login_date'].dt.strftime('%d-%b-%Y')
df = df[df['MIS_STATUS']=='Approved']
df1 = df.groupby(['ZONE','REGION'])['IN_Cr'].agg(count ="count",sum = 'sum')
df1 = df1.reset_index()
order = ['Assam','Bihar','Delhi','Haryana',
'Rajasthan'
,'Uttarakhand'
,'UP East'
,'UP West'
,'Varanasi'
,'AP 1'
,'AP 2'
,'Bangalore'
,'Mysore'
,'Telangana 1'
,'Telangana 2'
,'TN'
,'Mumbai','ROM','ROM 2','Gujarat','MP',]
df1['REGION'] = pd.Categorical(df1['REGION'],categories=order,ordered=True)
df1 = df1.sort_values('REGION').reset_index(drop = True)
east_total  = pd.DataFrame({'ZONE':['East'],
                            'REGION':['East Total'],
                            'count':[df1.loc[df1['REGION'].isin(['Assam','Bihar']),'count'].sum()],
                            'sum':[df1.loc[df1['REGION'].isin(['Assam','Bihar']),'sum'].sum()]})
pos = df1.loc[df1['ZONE']=='East'].index.max()+1
df1 = pd.concat([df1.iloc[:pos],east_total,df1.iloc[pos:]],ignore_index=True)
delhi_total = pd.DataFrame({'ZONE':['North'],'REGION':['Delhi Total'],
                            'count':[df1.loc[df1['REGION'].isin
                                             (['Delhi','Haryana','Rajasthan','Uttarakhand']),'count'].sum()],
                            'sum':[df1.loc[df1['REGION'].isin(['Delhi','Haryana','Rajasthan','Uttarakhand']),'sum'].sum()]})
utk_below = df1[df1['REGION']=="UTK"]
uttarakhand_below = df1[df1['REGION']=="Uttarakhand"]
delhi_below = df1[df1['REGION']=="Delhi"]
rj_below = df1[df1['REGION']=="Rajasthan"]
hy_below = df1[df1['REGION']=="Haryana"]

if not utk_below.empty:
    pos2 = utk_below.index.max()+1
elif not uttarakhand_below.empty:
    pos2 = uttarakhand_below.index.max()+1
elif not rj_below.empty:
    pos2 = rj_below.index.max()+1
elif not hy_below.empty:
    pos2 = hy_below.index.max()+1
elif not delhi_below.empty:
    pos2 = delhi_below.index.max()+1         

df1 = pd.concat([df1.iloc[:pos2],delhi_total,df1.iloc[pos2:]],ignore_index=True)
up_total = pd.DataFrame({'ZONE':['North'],'REGION':['UP Total'],
                         'count':[df1.loc[df1['REGION'].isin(['UP East','UP West','Varanasi']),'count'].sum()],
                         'sum':[df1.loc[df1['REGION'].isin(['UP East','UP West','Varanasi']),'sum'].sum()]})
vs_below = df1[df1['REGION']=='Varanasi']
west_below = df1[df1['REGION']=='UP West']
east_below = df1[df1['REGION']=='UP East']
if not vs_below.empty:
    pos3 = vs_below.index.max()+1
elif not west_below.empty:
    pos3 = west_below.index.max()+1
elif not east_below.empty:
    pos3 = east_below.index.max()+1        
df1 = pd.concat([df1.iloc[:pos3],up_total,df1.iloc[pos3:]],ignore_index=True)
north_total = pd.DataFrame({'ZONE':['North'],'REGION':['North Total'],
                            'count':[df1.loc[df1['REGION'].isin(['UP Total','Delhi Total']),'count'].sum()],
                            'sum':[df1.loc[df1['REGION'].isin(['UP Total','Delhi Total']),'sum'].sum()]})
up_below = df1[df1['REGION']=='UP Total']
delhi_below = df1[df1['REGION']=='Delhi Total']
if not up_below.empty:
    pos4 = up_below.index.max()+1
elif not delhi_below.empty:
    pos4 = delhi_below.index.max()+1    
df1 = pd.concat([df1.iloc[:pos4],north_total,df1.iloc[pos4:]],ignore_index=True)
ap_total = pd.DataFrame({'ZONE':['South'],'REGION':['AP Total'],
                         'count':[df1.loc[df1['REGION'].isin(['AP 1','AP 2']),'count'].sum()],
                         'sum':[df1.loc[df1['REGION'].isin(['AP 1','AP 2']),'sum'].sum()]})
ap2_below = df1[df1['REGION']=='AP 2']
ap1_below = df1[df1['REGION']=='AP 1']
 
if not ap2_below.empty:
    pos5 = ap2_below.index.max()+1
elif not ap1_below.empty:
    pos5 = ap1_below.index.max()+1    
df1 = pd.concat([df1.iloc[:pos5],ap_total,df1.iloc[pos5:]],ignore_index=True)
karnataka_total = pd.DataFrame({'ZONE':['South'],'REGION':['Karnataka Total'],
                                'count':[df1.loc[df1['REGION'].isin(['Bangalore','Mysore']),'count'].sum()],
                                'sum':[df1.loc[df1['REGION'].isin(['Bangalore','Mysore']),'sum'].sum()]})
pos6 = df1.loc[df1['REGION']=='Mysore'].index.max()+1
df1 = pd.concat([df1.iloc[:pos6],karnataka_total,df1.iloc[pos6:]],ignore_index=True)
telangana_total = pd.DataFrame({'ZONE':['South'],'REGION':['Telangana Total'],
                                'count':[df1.loc[df1['REGION'].isin(['Telangana 1','Telangana 2']),'count'].sum()],
                                'sum':[df1.loc[df1['REGION'].isin(['Telangana 1','Telangana 2']),'sum'].sum()]})
pos7 = df1.loc[df1['REGION']=='Telangana 2'].index.max()+1
df1 = pd.concat([df1.iloc[:pos7],telangana_total,df1.iloc[pos7:]],ignore_index=True)
south_total = pd.DataFrame({'ZONE':['South'],
                            'REGION':['South Total'],
                            'count':[df1.loc[df1['REGION'].isin(['Telangana Total','TN','Karnataka Total','AP Total']),'count'].sum()],
                            'sum':[df1.loc[df1['REGION'].isin(['Telangana Total','TN','Karnataka Total','AP Total']),'sum'].sum()]})
pos8 = df1.loc[df1['REGION']=='TN'].index.max()+1
df1 = pd.concat([df1.iloc[:pos8],south_total,df1.iloc[pos8:]],ignore_index = True)
mh_total = pd.DataFrame({'ZONE':['West'],'REGION':['Maharashtra Total'],
                         'count':[df1.loc[df1['REGION'].isin(['Mumbai','ROM','ROM 2']),'count'].sum()],
                         'sum':[df1.loc[df1['REGION'].isin(['Mumbai','ROM','ROM 2']),'sum'].sum()]})
pos9 = df1.loc[df1['REGION']=="ROM 2"].index.max()+1
df1 = pd.concat([df1.iloc[:pos9],mh_total,df1.iloc[pos9:]],ignore_index=True)
west_total = pd.DataFrame({'ZONE':['West'],'REGION':['West Total'],
                           'count':[df1.loc[df1['REGION'].isin(['Maharashtra Total','MP','Gujarat']),'count'].sum()],
                           'sum':[df1.loc[df1['REGION'].isin(['Maharashtra Total','MP','Gujarat']),'sum'].sum()]})
pos10 = df1.loc[df1['REGION']=='MP'].index.max()+1
df1 = pd.concat([df1.iloc[:pos10],west_total,df1.iloc[pos10:]])
Grand_Total = pd.DataFrame({'ZONE':['Total'],'REGION':['Grand Total'],
                            'count':[df1.loc[df1['REGION'].isin(['East Total','West Total','North Total','South Total']),'count'].sum()],
                            'sum':[df1.loc[df1['REGION'].isin(['East Total','West Total','North Total','South Total']),'sum'].sum()]})
df1 = pd.concat([df1,Grand_Total],ignore_index = True)
# pos11= df1.loc[len(df['REGION'])]
df1['sum'] = df1['sum'].round(2)
df1 = df1.drop(columns = ['ZONE'])

import os
path2 = r"C:\Users\ET0001301\Pictures\Data\Approved not Disb"
file_name = f'ME-MIS_Appoved Data {current_date}.xlsx' 
save_path = os.path.join(path2,file_name)
with pd.ExcelWriter(save_path,engine='openpyxl') as writer:
    df.to_excel(writer,sheet_name='Data',index=False)
    df1.to_excel(writer,sheet_name='Approved_Summary',index = False,startrow = 1)
    
# df1.to_excel(save_path,index = False,startrow = 1)
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side

wb = load_workbook(save_path)
ws = wb['Approved_Summary']
ws.merge_cells('A1:C1')
ws['A1']='Approved'
zone_total = ['East Total','West Total','North Total','South Total','Grand Total']
region_total = ['Delhi Total','UP Total','AP Total','Karnataka Total','Telangana Total','Maharashtra Total']
for rows in ws.iter_rows(min_row= 1,max_row= ws.max_row, min_col= 1,max_col = ws.max_column):
    for cell in rows:
        if cell.value and (cell.value in zone_total or cell.value == 'REGION'):
            for cell in rows:
                cell.fill = PatternFill(start_color='A9CCE3',fill_type='solid')
                cell.font=Font(bold = True)
for row in ws.iter_rows(min_row = 2,max_row = ws.max_row,min_col = 1,max_col = ws.max_column):
    for cell in row:
        if cell.value and cell.value in region_total:
            for cell in row:
                cell.fill = PatternFill(start_color='FADBD8',fill_type='solid')
                cell.font = Font(bold=True)
for row in ws.iter_rows(min_row = 1,max_row = ws.max_row,min_col = 1,max_col = ws.max_column):
    for cell in row:
        if cell.value:
            cell.alignment = Alignment(horizontal ='center',vertical = 'center')
            cell.border = Border(top = Side(border_style='thin'),bottom = Side(border_style='thin'),
                                 left = Side(border_style='thin'),
                                 right = Side(border_style='thin'))
for rows in ws[1]:
    rows.fill = PatternFill(start_color='A9CCE3',fill_type='solid')
    rows.font = Font(bold = True)
    
for cols in ws.iter_cols(min_row = 2,max_row = ws.max_row,min_col = 1,max_col = ws.max_column):
        max_length = 0
        col_letter = cols[0].column_letter
        for cell in cols:
            if cell.value is not None:
                max_length = max(max_length,len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length+2                        
wb.save(save_path)
wb.close()
html_table = df1.to_html(index=False)
outlook1 = win32.Dispatch('Outlook.Application')
mail = outlook1.CreateItem(0)
mail.To ='ravi.gupta@sbfc.com'
mail.subject = 'Approved but not Disbursed' 
mail_date = pd.Timestamp.now()
last_3_days2 = mail_date-timedelta(days=3)
last_3_days2 =last_3_days2.strftime('%d-%b') 
mail.HTMLBODY = f'''
<p>Dear All,</p>

<p>Please find attached list of cases that were Approved on or before {last_3_days2} but not yet Disbursed.</p>

<p>The Tranch Cases is also included.</p>

{html_table}

<p>Regrads,<br>
    Ravi Gupta</p>
'''
mail.Attachments.Add(save_path)
mail.Display()
mail.Send()
