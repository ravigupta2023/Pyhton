import pandas as pd
path = r"C:\Users\ET0001301\Pictures\Data\FCL\Ticket_Report_13-03-2026_After Tagging_open Tickets.xlsx"
df_ticket = pd.read_excel(path,sheet_name='Sheet0')

lap_data = r"C:\Users\ET0001301\Pictures\Data\FCL\1.LAP_PORTFOLIO_FEB26.xlsx"
df_lap = pd.read_excel(lap_data,sheet_name='Sheet1')

fcl_last_date = r"C:\Users\ET0001301\Pictures\Data\FCL\FCL Request 12th March'26.xlsx"
df_fcl = pd.read_excel(fcl_last_date,sheet_name='Fresh Cases')
df_ticket['Created Date'] = pd.to_datetime(df_ticket['Created Date'],errors = 'coerce')
df_ticket.info()
df = df_ticket[(df_ticket['Folder Level 2']=='FCL required') & (df_ticket['Created Date']>='23-02-2026')]
df['Loan Number'] = df['Loan Number'].str.strip()
df.insert(df.columns.get_loc('Loan Number')+1,'DPD',pd.NA)
df['DPD'] = df['Loan Number'].map(df_lap.set_index('PROSPECTCODE')['DPD_BKT']).fillna('#N/A')
df = df[df['DPD'].isin(['A.Current','B.1-30'])]

df = df[~df['Loan Number'].isin(df_fcl['Lan no.'])]
df= df.drop_duplicates(subset='Loan Number')
df = df.rename(columns={'Loan Number':'Lan no.'})
new_df = df['Lan no.']
new_df = new_df.to_frame(name = 'Lan no.')
new_df = new_df.merge(df_lap[['PROSPECTCODE','LEVIOSA ID','CUSTOMER_NAME','BRANCH','ME REGION','ZONE','POS IN Cr','DPD_BKT']],left_on='Lan no.',right_on='PROSPECTCODE',how = 'left')
new_df = new_df.drop(columns=['PROSPECTCODE'])
new_df['Created Date'] = new_df['Lan no.'].map(df.set_index('Lan no.')['Created Date'])
new_df['Month'] = new_df['Created Date'].apply(lambda x:x.strftime('%b') )
new_df = new_df.rename(columns = {'LEVIOSA ID':'Ref number','CUSTOMER_NAME':'Customer Name','BRANCH':'Branch','ME REGION':'Region 2','ZONE':'Zone','POS IN Cr':'New POS in cr','DPD_BKT':'DPD','Created Date':'In Date'})
df_fcl1 = pd.concat([df_fcl,new_df],ignore_index=True)
import numpy as np
df_fcl1['Month Number']= df_fcl1['In Date'].apply(lambda x:'Feb' if(x.month == 2) else x.day )
df_fcl1 = df_fcl1.drop(columns=['Retained Remark','Retained Stage','Branch remark ','Branch stage'])
df_fcl1 = df_fcl1.drop_duplicates(subset='Lan no.')
structure = ['Assam','Bihar','WB','Delhi','Haryana','Uttarakhand','PCH','Rajasthan','UP East','UP West','Varanasi','Bangalore 1','Bangalore 2','ROK','Telangana 1','Telangana 2','AP','TN','Gujarat','MP','Mumbai','ROM','ROM 2']
structure = pd.Series(structure,name = 'Region 2')
structure = structure.to_frame()
summary = df_fcl1.groupby(['Region 2','Month Number'])['New POS in cr'].sum().unstack(fill_value=0).reset_index()
structure = structure.merge(summary,on = 'Region 2',how = 'left')
cols = list(structure.columns)

# Keep first two columns fixed
fixed = ['Region 2']

# Other columns
others = [c for c in cols if c not in fixed]

# Put text first then numbers
text_cols = [c for c in others if isinstance(c, str)]
num_cols = [c for c in others if isinstance(c, (int, float))]

structure = structure[fixed + text_cols + num_cols]
type(df_fcl1['Month Number'])
max_num = df_fcl1['Month Number'].apply(pd.to_numeric,errors = 'coerce').max()
            #   numeric_only = True)
max_num = int(max_num)
east_data = {}
for i in range(1,max_num+1):
    east_value = structure.loc[structure['Region 2'].isin(['Assam','WB','Bihar']),i].sum()
    east_data[i]=east_value
east_total = pd.DataFrame({'Region 2':['East Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['Assam','WB','Bihar']),'Feb'].sum()],
                           **east_data})    
pos1 = structure[structure['Region 2']=='WB'].index.max()+1
structure = pd.concat([structure[:pos1],east_total,structure[pos1:]],ignore_index=True)
delhi_data = {}
for i in range(1,max_num+1):
    delhi_value = structure.loc[structure['Region 2'].isin(['Delhi','Haryana','Uttarakhand','PCH','Rajasthan']),i].sum()
    delhi_data[i]=delhi_value
delhi_total = pd.DataFrame({'Region 2':['Delhi Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['Delhi','Haryana','Uttarakhand','PCH','Rajasthan']),'Feb'].sum()],
                           **delhi_data})    
pos2 = structure.loc[structure['Region 2']=='Rajasthan'].index.max()+1
structure = pd.concat([structure[:pos2],delhi_total,structure[pos2:]],ignore_index=True)
up_data = {}
for i in range(1,max_num+1):
    up_value = structure.loc[structure['Region 2'].isin(['UP West','UP East','Varanasi']),i].sum()
    up_data[i]=up_value
up_total = pd.DataFrame({'Region 2':['UP Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['UP West','UP East','Varanasi']),'Feb'].sum()],
                           **up_data})    

pos3 = structure.loc[structure['Region 2']=='Varanasi'].index.max()+1
structure = pd.concat([structure[:pos3],up_total,structure[pos3:]],ignore_index=True)
north_data = {}
for i in range(1,max_num+1):
    north_value = structure.loc[structure['Region 2'].isin(['UP Total','Delhi Total']),i].sum()
    north_data[i]=north_value
north_total = pd.DataFrame({'Region 2':['North Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['UP Total','Delhi Total']),'Feb'].sum()],
                           **north_data})    

pos4 = structure.loc[structure['Region 2']=='UP Total'].index.max()+1
structure = pd.concat([structure[:pos4],north_total,structure[pos4:]],ignore_index=True)
karnataka_data = {}
for i in range(1,max_num+1):
    karnataka_value = structure.loc[structure['Region 2'].isin(['Bangalore 1','Bangalore 2','ROK']),i].sum()
    karnataka_data[i]=karnataka_value
karnataka_total = pd.DataFrame({'Region 2':['Karnataka Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['Bangalore 1','Bangalore 2','ROK']),'Feb'].sum()],
                           **karnataka_data})    

pos5 = structure.loc[structure['Region 2']=='ROK'].index.max()+1
structure = pd.concat([structure[:pos5],karnataka_total,structure[pos5:]],ignore_index=True)
telangana_data = {}
for i in range(1,max_num+1):
    telangana_value = structure.loc[structure['Region 2'].isin(['Telangana 1','Telangana 2']),i].sum()
    telangana_data[i]=telangana_value
telangana_total = pd.DataFrame({'Region 2':['Telangana Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['Telangana 1','Telangana 2']),'Feb'].sum()],
                           **telangana_data})    

pos6 = structure.loc[structure['Region 2']=='Telangana 2'].index.max()+1
structure = pd.concat([structure[:pos6],telangana_total,structure[pos6:]],ignore_index=True)
south_data = {}
for i in range(1,max_num+1):
    south_value = structure.loc[structure['Region 2'].isin(['TN','AP','Telangana Total','Karnataka Total']),i].sum()
    south_data[i]=south_value
south_total = pd.DataFrame({'Region 2':['South Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['TN','AP','Telangana Total','Karnataka Total']),'Feb'].sum()],
                           **south_data})    

pos7 = structure.loc[structure['Region 2']=='TN'].index.max()+1
structure = pd.concat([structure[:pos7],south_total,structure[pos7:]],ignore_index=True)
mh_data = {}
for i in range(1,max_num+1):
    mh_value = structure.loc[structure['Region 2'].isin(['Mumbai','ROM','ROM 2']),i].sum()
    mh_data[i]=mh_value
mh_total = pd.DataFrame({'Region 2':['Maharashtra Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['Mumbai','ROM','ROM 2']),'Feb'].sum()],
                           **mh_data})    

pos8 = structure.loc[structure['Region 2']=='ROM 2'].index.max()+1
structure = pd.concat([structure[:pos8],mh_total,structure[pos8:]],ignore_index=True)
west_data = {}
for i in range(1,max_num+1):
    west_value = structure.loc[structure['Region 2'].isin(['Maharashtra Total','Gujarat','MP']),i].sum()
    west_data[i]=west_value
west_total = pd.DataFrame({'Region 2':['West Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['Maharashtra Total','Gujarat','MP']),'Feb'].sum()],
                           **west_data})    

pos9 = structure.loc[structure['Region 2']=='Maharashtra Total'].index.max()+1
structure = pd.concat([structure[:pos9],west_total,structure[pos9:]],ignore_index=True)
grand_data = {}
for i in range(1,max_num+1):
    grand_value = structure.loc[structure['Region 2'].isin(['West Total','North Total','South Total','East Total']),i].sum()
    grand_data[i]=grand_value
grand_total = pd.DataFrame({'Region 2':['Grand Total'],
                           'Feb':[structure.loc[structure['Region 2'].isin(['West Total','North Total','South Total','East Total']),'Feb'].sum()],
                           **grand_data})    

pos10 = structure.loc[structure['Region 2']=='West Total'].index.max()+1
structure = pd.concat([structure[:pos10],grand_total,structure[pos10:]],ignore_index=True)
structure['Grand Total'] = structure.sum(axis = 1,numeric_only=True)
structure = structure.fillna(0)
from datetime import timedelta
today_date = pd.Timestamp.now()
yest_date = today_date-timedelta(days=1)
yest_date1 = yest_date.strftime('%d %b %y')

save_path = r"C:\Users\ET0001301\Pictures\Data\FCL"
os.makedirs(save_path,exist_ok=True)
file_name = f'FCL Request {yest_date1}.xlsx'
import os
final_path = os.path.join(save_path,file_name)

with pd.ExcelWriter(final_path,engine='openpyxl') as writer:
    
    # df.to_excel(writer,sheet_name="DATA",index=False)
    df_fcl1.to_excel(writer,sheet_name='Fresh Cases',index=False)
    structure.to_excel(writer,sheet_name='Summary',index = False,startrow=1)
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Side,Alignment,Border
wb = load_workbook(final_path)
ws = wb['Summary']
color = PatternFill(start_color='A9CCE3',fill_type='solid')
color2 = PatternFill(start_color='FCD5B4',fill_type='solid')
align = Alignment(horizontal='center',vertical='center')
bold = Font(bold=True)
bord = Side(style='thin')
bord2 = Border(left = bord,
               right= bord,
               top = bord,
               bottom=bord
               )
for cols in ws.iter_cols(min_row=3,min_col=2):
    for cells in cols:
        if  isinstance(cells.value,(int,float)):
            cells.number_format = '0.00'

for cols in ws.iter_cols():
    for cell in cols:
        cell.alignment = align
        cell.border = bord2
regions = ['Delhi Total','UP Total','Karnataka Total','Telangana Total','Maharashtra Total']
zones = ['West Total','North Total','East Total','South Total','Grand Total','Region 2']
for cols in ws.iter_rows():
    for cell in cols:
        if cell.value in regions:
            for cell in cols:
                cell.fill = color2
                cell.font = bold    
    for cell in cols:
        if cell.value in zones:
            for cell in cols:
                cell.fill = color
                cell.font = bold
for cols in ws.iter_cols():
    max_length = 0
    col_letter = cols[0].column_letter 
    for cell in cols:
        max_length = max(max_length,len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length                                  

ws.merge_cells(start_row=1,end_row=1,start_column=3,end_column=ws.max_column)
ws.cell(row=1, column=3).value = "March'26"
for cell in ws[1]:
    if cell.value:
        cell.font = bold
        cell.fill = color
        cell.border = bord2
        cell.alignment = align
wb.save(final_path)
wb.close()
