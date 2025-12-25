import pandas as pd
import os
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side,PatternFill,Alignment
import win32com.client as win32
import shutil

outlook = win32.Dispatch('Outlook.Application').GetNamespace('MAPI')
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items
messages.sort("[ReceivedTime]",True)

print("START FETCHING MAIL FROM OUTLOOK")

sub_current_date = pd.Timestamp.now().strftime('%d%b%Y')

outlook_path = r"C:\Users\et0001301\Desktop\Python\OUTLOOK DATA"
outlook_data_file_name = f'ME-MIS-SUMMARY-{sub_current_date}.xlsx'

os.makedirs(outlook_path,exist_ok = True)
outlook_final_path = os.path.join(outlook_path,outlook_data_file_name)

# sender details
source_subject = None
source_sender = None
source_received_time= None

for mail in messages:
    subject = mail.subject
    if 'ME-MIS-' in subject and sub_current_date in subject:
        if mail.Attachments.count>=2:
            second_attachment = mail.Attachments.Item(2)
            second_attachment.SaveAsFile(outlook_final_path)
        source_subject = mail.Subject
        source_sender = mail.SenderEmailAddress
        source_received_time = mail.ReceivedTime.strftime('%d-%b-%Y %H:%M:%S')   
        mail_info = {
                    "subject":mail.subject,
                    "sender":mail.sender,
                    'Received_time':mail.ReceivedTime
                    }       
        for k,v in mail_info.items():
            print(f'{k}:{v}')
        break         
print('MAIL GETTING FROM OUTLOOK SUCCESSFULL')

time.sleep(2)

print("MAIL SAVE SUCCESSFULL")

path = r"C:\Users\et0001301\Desktop\Python\MAIN DATA"
os.makedirs(path,exist_ok = True)
main_data_final_path = os.path.join(path,outlook_data_file_name)

shutil.copy(outlook_final_path,path)


df = pd.read_excel(main_data_final_path,sheet_name = 'out_file')

df.columns = df.columns.str.strip()

df.insert(12,'Decision_month',pd.NA)
df.insert(28,'Logins',pd.NA)

df['Tranch'] = df.apply(lambda x:'YES' if (
                                            x['SCHEMETYPE'] in {'BT Topup',
                                                'HL BT Top up',
                                                'Micro loans BT Top up',
                                                'SALARIED LAP BT TOPUP',
                                                'SBL_Top up',
                                                'Secured BL BT Top-up',
                                                'Top-up-PreApproved'
                                                })
                                        else "NO",
                                        axis = 1
                                        ) 

df['MIS_STATUS'] = df['MIS_STATUS'].apply(lambda x:'Approved' if x == 'APPROVED - ICICI' else ('Disbursed' if x=='DISBURSED - ICICI' else('WIP' if x == 'WIP - ICICI' else x)))

df['MIS_STATUS'] = df.apply(lambda x:'WIP' if ((x['CURRENTSTATUS'] in {'Hold','Queue'}) & (x['MIS_STATUS']=='Approved')) else x['MIS_STATUS'], axis = 1)

df['DOCUMENTRECEIVEDATCPADATE']= pd.to_datetime(df['DOCUMENTRECEIVEDATCPADATE'])
df['login_date'] = pd.to_datetime(df['login_date'])


current_date = pd.Timestamp.now()
current_year = current_date.year
current_month = current_date.month
current_month_name = current_date.strftime('%b')

df['Logins']=df['DOCUMENTRECEIVEDATCPADATE'].apply(lambda x:current_month_name if ((x.month == current_month) and (x.year == current_year)) else "")
df['Decision_month'] = df['login_date'].apply(lambda x:current_month_name if ((x.month == current_month) and (x.year == current_year)) else "")

df['login_date'] = df['login_date'].dt.strftime('%d-%b-%Y')
df['DOCUMENTRECEIVEDATCPADATE'] = df['DOCUMENTRECEIVEDATCPADATE'].dt.strftime('%d-%b-%Y')

df['REGISTRATIONStage'] = df['REGISTRATIONStage'].str.replace('KO','K0',case = False)

# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot
# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot
# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot
# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot# creating pivot

filter_for_logins = df[(df['Tranch'].str.lower()=="no") & (df['Logins'].str.lower()==current_month_name.lower())]
filter_for_approval = df[(df['Tranch'].str.lower()=='no') & (df['Decision_month'].str.lower()==current_month_name.lower()) & (df['MIS_STATUS'].isin(['Approved','Subjective Approval']))]
filter_for_disbursed = df[(df['Decision_month'].str.lower()==current_month_name.lower()) & (df['MIS_STATUS'].str.lower()=='Disbursed'.lower())]

print("DATA CLEANING DONE")
login_pivot = pd.pivot_table(
                        filter_for_logins,
                        index = 'REGISTRATIONStage',
                        columns = 'MIS_STATUS',
                        values = 'IN_Cr',
                        aggfunc = 'count',
                        fill_value = 0,
                        margins = True,
                        margins_name = 'Grand Total'
                        )

approval_pivot = pd.pivot_table(
                                filter_for_approval,
                                index = 'REGISTRATIONStage',
                                values = 'IN_Cr',
                                aggfunc = ['count','sum'],
                                fill_value = 0,
                                margins = True,
                                margins_name = 'Grand Total'
                                )
disbursal_pivot = pd.pivot_table(
                                filter_for_disbursed,
                                index = 'REGISTRATIONStage',
                                values = 'IN_Cr',
                                aggfunc = ['count','sum'],
                                fill_value = 0,
                                margins = True,
                                margins_name = 'Grand Total'
                                )
print("PIVOT CREATION DONE")

save_path = r"C:\Users\et0001301\Desktop\Python\MAIN DATA"
file_name = f'ME-MIS_Cleaned_data-{sub_current_date}.xlsx'
os.makedirs(save_path,exist_ok = True)
final_path = os.path.join(save_path,file_name)

with pd.ExcelWriter(final_path,engine = 'openpyxl') as writer:
    df.to_excel(writer,sheet_name = 'Data',index = False)
    
    tables = [
            ('Login_summary',login_pivot),
            ('App_summary',approval_pivot),
            ('Disbursal_summary',disbursal_pivot)
    ]
    
    # if 'pivots' not in writer.book.sheetnames:
    #     ws1 = writer.book.create_sheet('pivots')
    # else:
    #     ws1 = writer['pivots']    
    
    col = 0
    row = 0
    for titles,pivots in tables:
        ws1 = writer.book.create_sheet('pivots') if 'pivots' not in writer.book.sheetnames else writer.book['pivots']
        ws1.cell(row =row+1,column= col+1).value = titles
        pivots.to_excel(writer,sheet_name = 'pivots',startrow = row+1,startcol =col )
        col +=pivots.shape[1]+2

print("DATA AND PIVOT STORED IN THE FILE ")
print("NOW FORMATTING THE SHEET")
time.sleep(3)        
# loading Workbook
wb = load_workbook(final_path)
ws1 = wb['Data']
ws2 = wb['pivots']
print("DATA LOADED IN OPENPYXL FOR FOMATTING")
# defining the formatting
font_header = Font(bold = True,name = 'calibri')
font_value = Font(name = 'calibri')
header_color = PatternFill(start_color = '87CEEB',fill_type = 'solid')
pivot_header = PatternFill(start_color = '87CEEB',fill_type = 'solid')
border = Border(
                left = Side(style = 'thin'),
                right = Side(style = 'thin'),
                top = Side(style = 'thin'),
                bottom = Side(style = 'thin')
                )
alignment = Alignment(horizontal = 'center',vertical ='center')
print("FORMATTING VARIABLE CREATED")
print("FORMATTING THE DATA SHEET")
# formatting the data sheet
for row in ws1.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = alignment
        cell.font = font_value
for cell in ws1[1]:
    cell.fill = header_color
    cell.font = font_header
        
for col in ws1.iter_cols(min_row = 1,max_row = ws1.max_row,min_col=1,max_col = ws1.max_column):
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value is not None:
            max_length = max(max_length,len(str(cell.value)))
    ws1.column_dimensions[col_letter].width = max_length + 4 
print("DATA SHEET FORMATING DONE")    
# pivot formatting
print("NOW START FORMATTING THE PIVOT SHEET")
# formatting the first row that is the pivot name
for cell in ws2[1]:
    if cell.value is not None:  
        cell.fill = header_color
        cell.font = font_header
        cell.alignment = alignment
# highlighting the 2nd row that is column name 
for row in ws2.iter_rows(min_row = 2,max_row = 2,min_col = 1,max_col = 8 ):
    for cell in row:
        cell.fill = header_color
        cell.border = border
        cell.alignment = alignment
#  formatting the values of the first pivot
for row in ws2.iter_rows(min_row = 3,max_row = ws2.max_row,min_col = 1,max_col = 8):
    for cell in row:
        if cell.value is not None:
            cell.border = border
            cell.font = font_value
            cell.alignment = alignment
# merging the 2nd  and 3ed pivot headers 
ws2.merge_cells("K2:K3")
ws2.merge_cells("L2:L3")
ws2.merge_cells("O2:O3")
ws2.merge_cells("P2:P3")
# replacing hte name of the 2nd pivot and the 3rd pivot
ws2['K2']='Count'
ws2['L2']='sum'
ws2['O2']='count'
ws2['P2']='sum'

# formatting the values of the 2nd pivot
# fomatting the values of the 2nd pivot
for row in ws2.iter_rows(min_row = 2 ,max_row = ws2.max_row, min_col = 10,max_col = 12  ):
    for cell in row:
        if cell.value is not None:
            cell.border = border
            cell.alignment = alignment
            cell.font = font_value
# formatting the header  
for row in ws2.iter_rows(min_row = 2,max_row = 3,min_col = 10,max_col = 12):
    for cell in row:
        if cell.value is not None:
            cell.fill = header_color
            cell.font = font_header
# formatting the 3rd row
# formatting the header
for row in ws2.iter_rows(min_row = 2,max_row = 3,min_col = 14,max_col = 16):
    for cell in row:
        if cell.value is not None:
            cell.fill = header_color
            cell.font  = font_header
# formatting the whole pivot now
for row in ws2.iter_rows(min_row = 2,max_row = ws2.max_row,min_col = 14,max_col = 16):
    for cell in row:
        if cell.value is not None:
            cell.border = border
            cell.alignment = alignment
            cell.font = font_value
# formatting the whole pivot sheet.
#  Adjusting the column width 
for col in ws2.iter_cols(min_row = 1,max_row = ws2.max_row,min_col = 1 ,max_col = ws2.max_column):
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value is not None:
            max_length = max(max_length,len(str(cell.value)))
    ws2.column_dimensions[col_letter].width  = max_length  + 3                    
print("PIVOT SHEET FORMATTING DONE") 
wb.save(final_path) # saving the file  
wb.close() # closing the path   

print("WORKBOOK CLOSED")
print("WORKING FOR MAIL SENDING")
current_date_for_mail = pd.Timestamp.now().strftime('%d%b%Y')

send_outlook = win32.Dispatch('Outlook.Application')
mail = send_outlook.CreateItem(0)
mail.to = 'ravi.gupta@sbfc.com'
mail.Subject = f'ME-MIS_Summary-{current_date_for_mail}'
mail.HTMLBody = f"""
<html>
<body>
<p> 'Hello team,</p>

<p> Please find  attached ME-MIS-Summary as on {current_date_for_mail}.</p>

<p><b>SOURCE MAIL DETAILS:</b></p>
<table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;">
<tr>
    <td><b>Sender</b></td>
    <td>{source_sender}</td>
</tr>
<tr>
    <td><b>Subject</b></td>
    <td>{source_subject}</td>
</tr>
<tr>
    <td><b>Received Time</b></td>
    <td>{source_received_time}</td>
</tr>
</table>

<br>

<p> Regards,</p>

<p> Ravi Gupta</p>

</body>
</html>
"""

mail.Attachments.Add(final_path)
mail.Send() 


print("MAIL SEND SUCCESSFULL")        
        























