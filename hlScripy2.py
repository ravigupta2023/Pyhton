import win32com.client as win
import pythoncom
import os
import shutil
pythoncom.CoInitialize()
import pandas as pd
outlook = win.Dispatch('Outlook.Application').GetNameSpace('MAPI')
inbox = outlook.GetDefaultFolder(6)
messages = inbox.items
messages.Sort('[ReceivedTime]',True)
mail_date = pd.Timestamp.now().strftime('%d%b%Y')
Outlook_file_save_path = r"C:\Users\ET0001301\Pictures\Data\HL\Outlook data"
os.makedirs(Outlook_file_save_path,exist_ok=True)
file_name1 = f'ME HL Data {mail_date}.xlsx'
Outlook_file_save_path1 = os.path.join(Outlook_file_save_path,file_name1)
for mail in messages:
    sub = mail.subject
    if f'HL-MIS-{mail_date}' in sub:
        if mail.Attachments.Count>=2:
            second_attachment = mail.Attachments.Item(2)
            second_attachment.SaveAsFile(Outlook_file_save_path1)
        mail_info = {
            'sub':mail.subject,
            'sender':mail.sender,
            'ReceivedTime':mail.ReceivedTime
        }    
        for k,v in mail_info.items():
            print(k,v)
        break
                
shutil_path = r"C:\Users\ET0001301\Pictures\Data\Shutil Data"
os.makedirs(shutil_path,exist_ok=True)
file_name2 = file_name1 
shutil.copy(Outlook_file_save_path1,shutil_path)
path  = os.path.join(shutil_path,file_name1)
df = pd.read_excel(path,sheet_name = 'out_file')
df = df[df['REGION']=='Gujarat']
df['Tranch'] = df['SCHEMETYPE'].apply(lambda x:'Yes' if x=='HL BT Top up'
                                      else 'NO')
df['MIS_STATUS'] = df.apply(lambda x:'WIP' if ((x['CURRENTSTATUS'] in (['Hold','Queue'])) and 
                                               x['MIS_STATUS']=='Approved')
                            else x['MIS_STATUS'],axis= 1)
df['login_date'] = pd.to_datetime(df['login_date'],errors = 'coerce')
df['DOCUMENTRECEIVEDATCPADATE'] = pd.to_datetime(df['DOCUMENTRECEIVEDATCPADATE'],errors = 'coerce')

df.insert(df.columns.get_loc('login_date')+1,'Dec_month',pd.NA)
df.insert(df.columns.get_loc('DOCUMENTRECEIVEDATCPADATE')+1,'Logins',pd.NA)

current_date = pd.Timestamp.now()
current_month = pd.Timestamp.now().month
current_year = pd.Timestamp.now().year
current_month_name = pd.Timestamp.now().strftime('%b')

df['Dec_month'] = df['login_date'].apply(lambda x:current_month_name if ((x.month == current_month)
                                                                         and
                                                                         (x.year == current_year)and (pd.notna(x))) else "" )
df['Logins'] = df['DOCUMENTRECEIVEDATCPADATE'].apply(lambda x:current_month_name
                                                     if ((x.month == current_month)
                                                                         and
                                                                         (x.year == current_year)and (pd.notna(x)))
                                                     else ""
                                                     )

Branch = '''Anand
Gurukul
Maninagar
Mavdi
Mehsana
Morbi
Palanpur
Rajkot
Silvassa
Surat
HIMATNAGAR
Vadodara
Vapi
'''
Branch1 = Branch.split()
Branch2 = []
for cols in Branch1:
    Branch2.append(cols.upper())
Branch3 = pd.Series(Branch2,name = 'BRANCH')
Branch4 = Branch3.to_frame()
Logins_summary = df[((df['Tranch']=='NO') & (df['Logins'] != ""))]
Approved_summary = df[((df['Tranch']=='NO') & (df['Dec_month'] != "") & (df['MIS_STATUS']=='Approved'))]
Sub_app_summary = df[((df['Tranch']=='NO') & (df['Dec_month'] != "") & (df['MIS_STATUS']=='Subjective Approval'))]
WIP_summary = df[((df['Tranch']=='NO') & (df['Dec_month'] != "") & (df['MIS_STATUS']=='WIP'))]
Disb_summary = df[((df['Dec_month'] != "") & (df['MIS_STATUS']=='Disbursed'))]
Declined_summary = df[(df['Dec_month'] != "") & (df['MIS_STATUS']=='Declined')]

logins = Logins_summary.groupby('BRANCH')['IN_Cr'].agg( Login_Count = 'count',Login_Amt = 'sum')
Branch4 = Branch4.merge(logins,on = "BRANCH",how = 'left')
Approved = Approved_summary.groupby(['BRANCH'])['IN_Cr'].agg(Apprvoed_count = 'count',Approved_amt = 'sum')
Branch4 = Branch4.merge(Approved,on = 'BRANCH',how = 'left')
Sub_approved = Sub_app_summary.groupby(['BRANCH'])['IN_Cr'].agg(Sub_App_count = 'count',Sub_App_amt = 'sum')
Branch4 = Branch4.merge(Sub_approved,on = 'BRANCH',how = 'left')
wip = WIP_summary.groupby(['BRANCH'])['IN_Cr'].agg(WIP_count = 'count',WIP_amt = 'sum')
Branch4 = Branch4.merge(wip,on = 'BRANCH',how = 'left')
Disb = Disb_summary.groupby(['BRANCH'])['IN_Cr'].agg(Disb_count = 'count',Disb_amt = 'sum')
Branch4 = Branch4.merge(Disb,on = 'BRANCH',how = 'left')
Declined = Declined_summary.groupby(['BRANCH'])['IN_Cr'].agg(Declined_count = 'count',Declined_amt = 'sum')
Branch4 = Branch4.merge(Declined,on = 'BRANCH',how = 'left')
Branch4  = Branch4.fillna(0)
Branch4 = Branch4.round(2)
Branch4['Total Count'] = Branch4[['Apprvoed_count','Sub_App_count','Disb_count']].sum(axis=1)
Branch4['Total Amt'] = Branch4[['Approved_amt','Sub_App_amt','Disb_amt']].sum(axis=1)

total_row = pd.DataFrame({'BRANCH':['Total'],
                          'Login_Count':[Branch4['Login_Count'].sum()],
                          'Login_Amt':[Branch4['Login_Amt'].sum()],
                          'Apprvoed_count':[Branch4['Apprvoed_count'].sum()],
                          'Approved_amt':[Branch4['Approved_amt'].sum()],
                          'Sub_App_count':[Branch4['Sub_App_count'].sum()],
                          'Sub_App_amt':[Branch4['Sub_App_amt'].sum()],
                          'WIP_count':[Branch4['WIP_count'].sum()],
                          'WIP_amt':[Branch4['WIP_amt'].sum()],
                          'Disb_count':[Branch4['Disb_count'].sum()],
                          'Disb_amt':[Branch4['Disb_amt'].sum()],
                          'Declined_count':[Branch4['Declined_count'].sum()],
                          'Declined_amt':[Branch4['Declined_amt'].sum()],
                          'Total Count':[Branch4['Total Count'].sum()],
                          'Total Amt':[Branch4['Total Amt'].sum()],
                          
                          })
Branch4=pd.concat([Branch4,total_row],ignore_index=True)
# Branch4['Login_Amt'] = Branch4['Login_Amt'].round(2)
# Branch4['Amt'] = Branch4['Amt'].map('{:.2f}'.format)
Branch4
Branch4.rename(columns = {'Login_Count':'Count','Login_Amt':'Amt',
                                            'Apprvoed_count':'Count','Approved_amt':'Amt',
                                            'Sub_App_count':'Count','Sub_App_amt':'Amt',
                                            'WIP_count':'Count','WIP_amt':'Amt',
                                            'Disb_count':'Count','Disb_amt':'Amt',
                                            'Declined_count':'Count','Declined_amt':'Amt',
                                            'Total Count':'Count','Total Amt':'Amt'},inplace=True)
Branch4['BRANCH'] = Branch4['BRANCH'].str.capitalize()

from datetime import timedelta
save_file_date = pd.Timestamp.now()
yest_date = (save_file_date-timedelta(days =1)).strftime('%d %b %Y')
save_path = r"C:\Users\ET0001301\Pictures\Data\HL"
file_name = f'HL_Summary{yest_date}.xlsx'
import os
save_path2 = os.path.join(save_path,file_name)
with pd.ExcelWriter(save_path2,engine = 'openpyxl') as writer:
    df.to_excel(writer,sheet_name = 'Data',index = False)
    Branch4.to_excel(writer,sheet_name = 'Summary',index = False,startrow = 1)
    
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font,Side,Border,PatternFill
wb = load_workbook(save_path2)
ws = wb['Summary']
cols3 = [3,5,7,9,11,13,15]
for cols in ws.iter_cols(min_row=2):
    header = cols[0].value
    if header and 'Amt' in str(header):
        for cell in cols:
            if isinstance(cell.value,(int, float)):
                cell.number_format = '0.00'
merge_range = ('A1:A2','B1:C1','D1:E1','F1:G1','H1:I1','J1:K1','L1:M1','N1:O1')
for cells in merge_range:
    ws.merge_cells(cells)
ws['A1'] = 'Branch'
ws['B1'] = 'Logins'
ws['D1'] = 'Approved'
ws['F1']='Subjective Approval'
ws['H1'] = 'WIP'
ws['J1']='Disbused'
ws['L1']='Declined'
ws['N1'] = 'Total'

color = PatternFill(start_color='A9CCE3',fill_type='solid')
fonts = Font(bold=True)
Align = Alignment(horizontal='center',vertical = 'center')
bord = Border(left=Side(border_style='thin'),
              right=Side(border_style='thin'),
              bottom=Side(border_style='thin'),
              top=Side(border_style='thin'))
for cell in ws[1]:
    cell.alignment = Align
    cell.font =fonts
    cell.fill = color
    cell.border = bord
for cell in ws[2]:
    cell.alignment = Align
    cell.font =fonts
    cell.fill = color
    cell.border = bord
for cells in ws.iter_rows(min_row = 1,max_row = ws.max_row,min_col=1,max_col=ws.max_column):
    for cols in cells:
        cols.border = bord
        cols.alignment = Align
        if cols.value == 'Total':
            for cell in cells:
                cell.fill = color
                cell.font = fonts
            

        
wb.save(save_path2)
wb.close()

# html_table = Branch4.to_html(index=False)


html_table = Branch4.to_html(index=False)




outlook1 = win.Dispatch('Outlook.Application')
mail = outlook1.CreateItem(0)
mail.to = 'ravi.gupta@sbfc.com'
mail.subject = f'ME HL Gujarat Data as on {yest_date}'
mail.HTMLBODY = f'''
<p> Dear All,</p>
                                                                 
<p>Please find attached HL summary for Gujarat, branch wise.</p>
              
<p>The details include Logins | Approved | Subjective Approvals |WIP | Disbursed | Declined.</p>

{html_table}

<p>Regards,<br>Ravi Gupta</p>
'''
mail.Attachments.Add(save_path2)
mail.Display()
mail.Send()


