import pandas as pd
from openpyxl import load_workbook
import os

path = r"C:\Users\Ravi Gupta\Downloads\ME-MIS_HL_SUMMARY.xlsx"

df = pd.read_excel(path,sheet_name= 'out_file')

df = df[df['REGION']=='Gujarat']

df.loc[df['SCHEMETYPE']=='HL BT Top up','Tranch']='yes'

df['login_date']=pd.to_datetime(df['login_date'])

df['Decision_month']=df['login_date'].dt.month_name().apply(lambda x:pd.Timestamp.now().strftime('%b') if x==pd.Timestamp.now().strftime('%B') else "")

df['DOCUMENTRECEIVEDATCPADATE']=pd.to_datetime(df['DOCUMENTRECEIVEDATCPADATE'])

df['logins']=df['DOCUMENTRECEIVEDATCPADATE'].dt.month_name().apply(lambda x:pd.Timestamp.now().strftime('%b') if x == pd.Timestamp.now().strftime('%B') else "")

current_month_name = pd.Timestamp.now()
short_month_name = current_month_name.strftime('%b')

filter_for_logins = df[(df['logins']== short_month_name) & (df['Tranch']=='NO')]

filter_for_approval = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='Approved') & (df['Tranch']=='NO')]

filter_for_sub_approval = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='Subjective Approval') & (df['Tranch']=='NO')]

filter_for_wip = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='WIP') & (df['Tranch']=='NO')]

filter_for_disbursed = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='Disbursed') & (df['Tranch']=='NO')]

filter_for_declined = df[(df['Decision_month']==short_month_name) & (df['MIS_STATUS']=='Declined') & (df['Tranch']=='NO')]


def create_pivot(filtered_data):
    pivot = pd.pivot_table(filtered_data,
                           values = 'IN_Cr',
                           index = 'BRANCH',
                           aggfunc = ['count','sum'],
                           fill_value = 0,
                           margins = True,
                           margins_name = 'Total'
                           )
    return pivot
    
login_pivot = create_pivot(filter_for_logins)
approval_pivot = create_pivot(filter_for_approval)
sub_approval_pivot = create_pivot(filter_for_sub_approval)    
wip_pivot = create_pivot(filter_for_wip)
disbursed_pivot = create_pivot(filter_for_disbursed)
declined_pivot = create_pivot(filter_for_declined)

save_path = r"C:\Users\Ravi Gupta\OneDrive\Desktop\New folder"
os.makedirs(save_path,exist_ok = True )
file_name = 'Cleaned_data.xlsx'

final_path = os.path.join(save_path,file_name) 



with pd.ExcelWriter(final_path,engine = 'openpyxl',mode = 'w') as writer:
    df.to_excel(writer,sheet_name = 'cleaned_data',index = False)
    
    tables = [login_pivot,approval_pivot,sub_approval_pivot,wip_pivot,disbursed_pivot,declined_pivot]
            
    row = 0
    col = 0 
    for pivots in tables:
        pivots.to_excel(writer,sheet_name = 'pivots', startrow = row, startcol = col )
        col += pivots.shape[1]+2
        
from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side,Alignment,PatternFill

wb = load_workbook(final_path)
ws = wb['pivots']

font = Font(bold = True)
Alignments1 = Alignment(horizontal = 'center',vertical = 'center')
Header_Fill = PatternFill(start_color = 'ADD8E6',fill_type = 'solid')
total_fill = PatternFill(start_color = 'FFFF00',fill_type = 'solid')
thin_border = Border(left = Side(style = 'thin'),
                     right = Side(style = 'thin'),
                     top = Side(style = 'thin'),
                     bottom = Side(style = 'thin'))
# total_fill = PatternFill(start_color = '')

for cell in ws[1]:
    if cell.value is not None:
        cell.font = font
        cell.fill = Header_Fill
 
 
for row in ws.iter_rows():
    for cell in row:
        if cell.value == 'Total':
            cell.fill = total_fill
    
    
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            cell.border = thin_border         
            cell.alignment = Alignments1

for col in ws.iter_cols():
    max_length = 0
    column_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length,len(str(cell.value)))
    ws.column_dimensions[column_letter].width= max_length + 3
wb.save(final_path)   

wb1 = load_workbook(final_path) 
ws1 = wb1['cleaned_data']

for col in ws1.iter_cols():
    max_length = 0
    column_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length,len(str(cell.value)))
    ws1.column_dimensions[column_letter].width = max_length + 4 
    
for rows in ws1.iter_rows():
    for cell in rows:
        cell.border = thin_border
        cell.alignment = Alignments1
 
for rows in ws1[1]:
    rows.font = font            
    

wb1.save(final_path)
            
        
print("save successfull")   